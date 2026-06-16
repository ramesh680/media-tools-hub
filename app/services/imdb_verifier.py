from __future__ import annotations

from datetime import date, datetime
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Callable
import csv
import re

from openpyxl import load_workbook

from app.models import utc_now_iso
from app.services.imdb import IMDbEnrichmentService, normalize_title


ProgressCallback = Callable[[int, str], None]

IMDB_VERIFIER_COLUMNS = [
    "Input Name",
    "Input Type",
    "Input Year",
    "Provided Code",
    "Matched Code",
    "Code Type",
    "Matched Name",
    "Title Type",
    "Start/Birth Year",
    "Genres/Profession",
    "Known For",
    "Match Status",
    "Lookup Note",
]


class IMDbBulkVerifierService:
    def __init__(self, imdb_service: IMDbEnrichmentService) -> None:
        self.imdb_service = imdb_service

    def verify_bulk(
        self,
        bulk_text: str = "",
        file_content: bytes | None = None,
        filename: str = "",
        progress: ProgressCallback | None = None,
    ) -> dict[str, Any]:
        if progress:
            progress(8, "Reading IMDb verification input")
        input_rows = self._parse_inputs(bulk_text, file_content, filename)
        if not input_rows:
            raise ValueError("Add at least one title/person row or upload a CSV/XLSX file.")

        if progress:
            progress(18, "Preparing local IMDb title/name datasets")
        self.imdb_service.ensure_index(progress=progress)

        output_rows: list[dict[str, Any]] = []
        matched = 0
        mismatched = 0
        not_found = 0
        total = max(len(input_rows), 1)
        for index, item in enumerate(input_rows, start=1):
            if progress:
                progress(62 + int((index / total) * 30), f"Checking {item.name}")
            output = self._verify_one(item)
            status = output["Match Status"]
            if status == "Matched":
                matched += 1
            elif status == "Code mismatch":
                mismatched += 1
            elif status == "No match":
                not_found += 1
            output_rows.append(output)

        summary = (
            f"Verified {len(input_rows)} bulk IMDb rows using local IMDb title.basics and name.basics indexes. "
            f"Matched: {matched}. Code mismatches: {mismatched}. No match: {not_found}."
        )
        return {
            "tracker_type": "imdb_verifier",
            "title": "IMDb Bulk Verification",
            "created_at": utc_now_iso(),
            "source_url": "https://datasets.imdbws.com/",
            "summary": summary,
            "sections": [
                {
                    "key": "imdb_verifier",
                    "title": "IMDb Verification Results",
                    "columns": IMDB_VERIFIER_COLUMNS,
                    "rows": output_rows,
                    "row_count": len(output_rows),
                    "supports_google": False,
                }
            ],
        }

    def _parse_inputs(
        self,
        bulk_text: str,
        file_content: bytes | None,
        filename: str,
    ) -> list["IMDbVerifyInput"]:
        rows: list[IMDbVerifyInput] = []
        if file_content and filename:
            suffix = Path(filename).suffix.lower()
            if suffix == ".csv":
                rows.extend(_inputs_from_csv(file_content))
            elif suffix in {".xlsx", ".xlsm"}:
                rows.extend(_inputs_from_workbook(file_content))
            else:
                raise ValueError("Upload a CSV, XLSX, or XLSM file for IMDb verification.")
        rows.extend(_inputs_from_text(bulk_text))
        return [row for row in rows if row.name]

    def _verify_one(self, item: "IMDbVerifyInput") -> dict[str, Any]:
        provided_code = _extract_any_imdb_code(item.provided_code)
        input_kind = _input_kind(item.input_type, provided_code)
        if input_kind == "name":
            return self._verify_name(item, provided_code)
        return self._verify_title(item, provided_code)

    def _verify_title(self, item: "IMDbVerifyInput", provided_code: str) -> dict[str, Any]:
        category = _category_for_title_lookup(item.input_type)
        match = self.imdb_service.lookup_title(
            item.name,
            category=category,
            release_date=_date_from_year(item.year),
        )
        if not match:
            return _output_row(item, provided_code, "", "ttcode", "", "", "", "", "", "No match", "No exact normalized title match in title.basics.")
        matched_code = match.get("imdb_id", "")
        status = _status_for_codes(provided_code, matched_code)
        note = "Matched by normalized title lookup in title.basics."
        if provided_code and provided_code != matched_code:
            note = f"Provided code differs from title.basics lookup. Suggested ttcode: {matched_code}."
        return _output_row(
            item,
            provided_code,
            matched_code,
            "ttcode",
            match.get("title", ""),
            match.get("title_type", ""),
            match.get("start_year", ""),
            match.get("genres", ""),
            "",
            status,
            note,
        )

    def _verify_name(self, item: "IMDbVerifyInput", provided_code: str) -> dict[str, Any]:
        match = self.imdb_service.lookup_name(item.name, profession_hint=item.input_type)
        if not match:
            return _output_row(item, provided_code, "", "nmcode", "", "", "", "", "", "No match", "No exact normalized person match in name.basics.")
        matched_code = match.get("imdb_id", "")
        status = _status_for_codes(provided_code, matched_code)
        note = "Matched by normalized person lookup in name.basics."
        if provided_code and provided_code != matched_code:
            note = f"Provided code differs from name.basics lookup. Suggested nmcode: {matched_code}."
        return _output_row(
            item,
            provided_code,
            matched_code,
            "nmcode",
            match.get("name", ""),
            "",
            match.get("birth_year", ""),
            match.get("primary_profession", ""),
            match.get("known_for_titles", ""),
            status,
            note,
        )


class IMDbVerifyInput:
    def __init__(self, name: str, input_type: str = "", year: str = "", provided_code: str = "") -> None:
        self.name = _clean(name)
        self.input_type = _clean(input_type)
        self.year = _clean(year)
        self.provided_code = _clean(provided_code)


def _inputs_from_text(raw_text: str) -> list[IMDbVerifyInput]:
    rows = []
    for line in raw_text.splitlines():
        cleaned = line.strip()
        if not cleaned:
            continue
        parts = [part.strip() for part in cleaned.split("|")]
        rows.append(
            IMDbVerifyInput(
                name=parts[0] if len(parts) >= 1 else "",
                input_type=parts[1] if len(parts) >= 2 else "",
                year=parts[2] if len(parts) >= 3 else "",
                provided_code=parts[3] if len(parts) >= 4 else "",
            )
        )
    return rows


def _inputs_from_csv(content: bytes) -> list[IMDbVerifyInput]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(StringIO(text))
    return [_input_from_mapping(row) for row in reader]


def _inputs_from_workbook(content: bytes) -> list[IMDbVerifyInput]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook.worksheets[0]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_normalize_header(value) for value in rows[0]]
    output = []
    for values in rows[1:]:
        if not any(_clean(value) for value in values):
            continue
        row = {headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))}
        output.append(_input_from_mapping(row))
    return output


def _input_from_mapping(row: dict[Any, Any]) -> IMDbVerifyInput:
    return IMDbVerifyInput(
        name=_first_value(row, ["name", "title", "title name", "input name", "query"]),
        input_type=_first_value(row, ["type", "input type", "title category", "title_category", "category"]),
        year=_first_value(row, ["year", "input year", "release year", "start year", "release date", "release_date"]),
        provided_code=_first_value(row, ["imdb id", "imdb_id", "imdb url", "imdb_url", "ttcode", "nmcode", "provided code"]),
    )


def _first_value(row: dict[Any, Any], keys: list[str]) -> str:
    normalized = {_normalize_header(key): value for key, value in row.items()}
    for key in keys:
        value = normalized.get(_normalize_header(key))
        if _clean(value):
            return _clean(value)
    return ""


def _input_kind(input_type: str, provided_code: str) -> str:
    normalized_type = normalize_title(input_type)
    if provided_code.startswith("nm"):
        return "name"
    if normalized_type in {"person", "people", "name", "talent", "actor", "actress", "director", "producer"}:
        return "name"
    return "title"


def _category_for_title_lookup(input_type: str) -> str:
    normalized = normalize_title(input_type)
    if normalized in {"tv", "tv show", "tv shows", "tv series", "series"}:
        return "TV Shows"
    return "Movies"


def _date_from_year(value: str) -> date | None:
    text = _clean(value)
    match = re.search(r"\b(19\d{2}|20\d{2})\b", text)
    if not match:
        return None
    try:
        return date(int(match.group(1)), 1, 1)
    except ValueError:
        return None


def _extract_any_imdb_code(value: str) -> str:
    text = _clean(value)
    match = re.search(r"\b(?:tt|nm)\d{7,12}\b", text, flags=re.IGNORECASE)
    return match.group(0).lower() if match else ""


def _status_for_codes(provided_code: str, matched_code: str) -> str:
    if not matched_code:
        return "No match"
    if provided_code and provided_code != matched_code:
        return "Code mismatch"
    return "Matched"


def _output_row(
    item: IMDbVerifyInput,
    provided_code: str,
    matched_code: str,
    code_type: str,
    matched_name: str,
    title_type: str,
    year: str,
    genres_or_profession: str,
    known_for: str,
    status: str,
    note: str,
) -> dict[str, Any]:
    return {
        "Input Name": item.name,
        "Input Type": item.input_type,
        "Input Year": item.year,
        "Provided Code": provided_code,
        "Matched Code": matched_code,
        "Code Type": code_type,
        "Matched Name": matched_name,
        "Title Type": title_type,
        "Start/Birth Year": year,
        "Genres/Profession": genres_or_profession,
        "Known For": known_for,
        "Match Status": status,
        "Lookup Note": note,
    }


def _normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("_", " ").strip().lower())


def _clean(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()
