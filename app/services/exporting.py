from __future__ import annotations

from io import BytesIO, StringIO
from pathlib import Path
from typing import Any
import csv
import json
import os
import uuid

from fastapi.responses import HTMLResponse, RedirectResponse, Response, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models import ExportPayload, safe_filename
from app.services.cache import TTLCache


class ExportService:
    def __init__(self, ttl_seconds: int, google_credentials: str = "") -> None:
        self.cache: TTLCache[ExportPayload] = TTLCache(ttl_seconds)
        self.google_credentials = google_credentials

    def register(self, payload: ExportPayload) -> str:
        export_id = uuid.uuid4().hex
        self.cache.set(export_id, payload)
        return export_id

    def register_snapshot_exports(self, snapshot: dict[str, Any]) -> dict[str, Any]:
        # Tag the run with its ingest title category (blank when the tool's rows
        # don't map to one), so the result panel can offer the second download
        # option: the same data as an ingest template.
        try:
            from app.services.ingest_export import ingest_category_for_tracker
            snapshot["ingest_category"] = ingest_category_for_tracker(
                snapshot.get("tracker_type", ""))
        except Exception:  # never let this break an export
            snapshot["ingest_category"] = ""
        for section in snapshot.get("sections", []):
            payload = ExportPayload(
                title=section["title"],
                columns=section["columns"],
                rows=section["rows"],
                tracker_type=snapshot["tracker_type"],
                section_key=section["key"],
                supports_google=bool(section.get("supports_google")),
            )
            section["export_id"] = self.register(payload)
        return snapshot

    def render_export(self, export_id: str, fmt: str) -> Response:
        payload = self.cache.get(export_id)
        fmt = fmt.lower()
        if payload is None:
            return HTMLResponse(
                "<main class='export-error'><h1>Export expired.</h1>"
                "<p>Reopen the run from Recent tracker runs to recreate export links.</p></main>",
                status_code=410,
            )
        if fmt == "csv":
            return self._csv_response(payload)
        if fmt in {"xlsx", "excel"}:
            return self._xlsx_response(payload)
        if fmt in {"google", "sheets", "gsheets"}:
            return self._google_sheets_response(payload)
        if fmt == "ingest":
            return self._ingest_response(payload)
        return HTMLResponse(
            "<main class='export-error'><h1>Unsupported export format.</h1>"
            "<p>Use CSV, Excel, Google Sheets, or Ingest Template.</p></main>",
            status_code=404,
        )

    def _ingest_response(self, payload: ExportPayload) -> Response:
        """Second download option: the SAME run's data turned into the ingest
        template for that tool's title category (TV Shows / Movies / Talent /
        Video Game), using the Title Automation ingest-template logic."""
        from app.services.bdr_ingest import BdrIngestError
        from app.services.ingest_export import (convert_rows_to_ingest,
                                                ingest_slug_for_tracker)

        slug = ingest_slug_for_tracker(payload.tracker_type)
        if not slug:
            return HTMLResponse(
                "<main class='export-error'><h1>No ingest template for this tool.</h1>"
                "<p>This report's rows do not map to a single title category, so an "
                "ingest template cannot be generated from it.</p></main>",
                status_code=404,
            )
        rows = [dict(zip(payload.columns, r)) if isinstance(r, (list, tuple)) else dict(r)
                for r in payload.rows]
        try:
            result = convert_rows_to_ingest(rows, slug)
        except BdrIngestError as exc:
            return HTMLResponse(
                "<main class='export-error'><h1>Could not build the ingest template.</h1>"
                f"<p>{exc}</p></main>",
                status_code=502,
            )
        return StreamingResponse(
            BytesIO(result["content"]),
            media_type=result["media_type"],
            headers={"Content-Disposition": f'attachment; filename="{result["filename"]}"'},
        )

    def _csv_response(self, payload: ExportPayload) -> StreamingResponse:
        buffer = StringIO()
        writer = csv.DictWriter(buffer, fieldnames=payload.columns, extrasaction="ignore", lineterminator="\n")
        writer.writeheader()
        for row in payload.rows:
            writer.writerow({column: row.get(column, "") for column in payload.columns})
        content = buffer.getvalue().encode("utf-8-sig")
        headers = {"Content-Disposition": f'attachment; filename="{safe_filename(payload.title, "csv")}"'}
        return StreamingResponse(iter([content]), media_type="text/csv; charset=utf-8", headers=headers)

    def _xlsx_response(self, payload: ExportPayload) -> StreamingResponse:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Export"
        worksheet.append(payload.columns)

        header_fill = PatternFill("solid", fgColor="EAF0F6")
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

        for row in payload.rows:
            worksheet.append([row.get(column, "") for column in payload.columns])

        worksheet.freeze_panes = "A2"
        for index, column in enumerate(payload.columns, start=1):
            max_width = len(column)
            for cell in worksheet[get_column_letter(index)]:
                max_width = max(max_width, len(str(cell.value or "")))
            worksheet.column_dimensions[get_column_letter(index)].width = min(max_width + 2, 60)

        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        headers = {"Content-Disposition": f'attachment; filename="{safe_filename(payload.title, "xlsx")}"'}
        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )

    def _google_sheets_response(self, payload: ExportPayload) -> Response:
        if not payload.supports_google:
            return HTMLResponse(
                "<main class='export-error'><h1>Google Sheets export is not enabled for this tracker.</h1>"
                "<p>Use CSV or Excel for this result.</p></main>",
                status_code=400,
            )

        credentials_source = self.google_credentials or os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON") or os.getenv(
            "GOOGLE_APPLICATION_CREDENTIALS"
        )
        if not credentials_source:
            return HTMLResponse(
                "<main class='export-error'><h1>Google Sheets credentials are missing.</h1>"
                "<p>Set GOOGLE_APPLICATION_CREDENTIALS to a service-account JSON file, or set "
                "GOOGLE_SERVICE_ACCOUNT_JSON to a JSON payload or file path, then retry the export.</p></main>",
                status_code=500,
            )

        try:
            sheet_url = self._create_google_sheet(payload, credentials_source)
        except ImportError as exc:
            return HTMLResponse(
                "<main class='export-error'><h1>Google Sheets dependencies are unavailable.</h1>"
                f"<p>{exc}</p></main>",
                status_code=500,
            )
        except Exception as exc:
            return HTMLResponse(
                "<main class='export-error'><h1>Google Sheets export failed.</h1>"
                f"<p>{str(exc)}</p></main>",
                status_code=500,
            )
        return RedirectResponse(sheet_url, status_code=303)

    def _create_google_sheet(self, payload: ExportPayload, credentials_source: str) -> str:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build

        scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive.file"]
        source_path = None if credentials_source.lstrip().startswith("{") else Path(credentials_source)
        if source_path and source_path.exists():
            credentials = service_account.Credentials.from_service_account_file(str(source_path), scopes=scopes)
        else:
            credentials_info = json.loads(credentials_source)
            credentials = service_account.Credentials.from_service_account_info(credentials_info, scopes=scopes)

        service = build("sheets", "v4", credentials=credentials)
        spreadsheet = (
            service.spreadsheets()
            .create(body={"properties": {"title": payload.title}}, fields="spreadsheetId,spreadsheetUrl")
            .execute()
        )
        values = [payload.columns]
        values.extend([[row.get(column, "") for column in payload.columns] for row in payload.rows])
        service.spreadsheets().values().update(
            spreadsheetId=spreadsheet["spreadsheetId"],
            range="A1",
            valueInputOption="RAW",
            body={"values": values},
        ).execute()
        return spreadsheet["spreadsheetUrl"]
