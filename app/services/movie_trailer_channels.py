"""Movie Trailer Channel Validator.

Finds the official **distributor** YouTube channel for a movie by starting from
the trailer MovieInsider already knows about, instead of guessing from a YouTube
search. Then validates the team's existing channel mapping against it.

Why this exists alongside YouTubeReleaseVerifierService
------------------------------------------------------
The existing verifier calls ``search.list`` once per title. That costs
**100 quota units per title**, so a 10,000 unit/day project quota tops out
around 99 titles per day - and it has to guess which of the top 10 search
results is really the official trailer.

MovieInsider publishes the trailer's YouTube video ID directly in its thumbnail
URLs::

    https://s.movieinsider.com/images/clayface/ytimg/OGO4Mqvo3jI/hqdefault_m1784763184.jpg
                                                  └─ YouTube video ID ─┘

Feeding that into ``videos.list`` (1 unit per 50 videos) returns the uploader's
channel exactly, with no guessing:

===========================  ==================  ====================
Approach                     Cost per title      Titles per 10k quota
===========================  ==================  ====================
search.list (old verifier)   ~101 units          ~99
videos.list (this service)   ~0.02 units         ~500,000
===========================  ==================  ====================

Lookup tiers
------------
Each tier is bounded so a run stays fast and needs no persistence (Render's free
tier has no persistent disk and sleeps after ~15 min idle).

A. **Trailer listing sweep** - newest ``page_budget`` pages of /movie-trailers
   (24 cards each). Yields the trailer video ID straight away. Covers current
   and upcoming releases, which is the common case.
B. **A-Z browse** - for titles tier A missed, walk ``/alpha/<letter>`` (20 per
   page, ordered newest-first) to find the numeric movie ID, then read the
   trailer embed off ``/m<id>/<slug>``. Needed because MovieInsider routes on
   the numeric ID and **ignores the slug** - ``/m1/clayface`` serves Spy Kids 2 -
   so a movie page cannot be reached from a title alone.
C. **Existing search-based verifier** - final fallback for anything MovieInsider
   has no trailer for, so coverage is never worse than before.
"""
from __future__ import annotations

from html import unescape
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Callable, Iterable
import csv
import re

from openpyxl import load_workbook

from app.models import MOVIE_TRAILER_CHANNEL_COLUMNS, utc_now_iso
from app.services.http_client import HttpClient
from app.services.imdb import normalize_title


ProgressCallback = Callable[[int, str], None]

MOVIEINSIDER_BASE = "https://www.movieinsider.com"
TRAILERS_PATH = "/movie-trailers"
LISTING_PAGE_SIZE = 24
ALPHA_PAGE_SIZE = 20

YOUTUBE_VIDEOS_URL = "https://www.googleapis.com/youtube/v3/videos"
YOUTUBE_CHANNELS_URL = "https://www.googleapis.com/youtube/v3/channels"

# Bounded work per run. Tier A costs one request per page.
DEFAULT_PAGE_BUDGET = 25          # 25 * 24 = 600 newest trailers
DEFAULT_ALPHA_PAGE_BUDGET = 6     # 6 * 20 = 120 newest titles per letter

# Validation statuses
MATCH = "Match"
MISMATCH = "Mismatch"
NEEDS_REVIEW = "Needs review"
NOT_OFFICIAL_UPLOAD = "Not official upload"
NO_CHANNEL_PROVIDED = "No channel provided"
TITLE_NOT_FOUND = "Title not found"
NO_CHANNEL_FOUND = "No channel found"

THUMB_RE = re.compile(r"/images/(?P<slug>[^/]+)/ytimg/(?P<video_id>[A-Za-z0-9_-]{11})/", re.IGNORECASE)
DETAIL_RE = re.compile(r"/m(?P<movie_id>\d+)/(?P<slug>[^/\"'?#]+)/videos/(?P<video_page_id>\d+)", re.IGNORECASE)
MOVIE_LINK_RE = re.compile(r"/m(?P<movie_id>\d+)/(?P<slug>[a-z0-9\-]+)(?![\w/])", re.IGNORECASE)
EMBED_RE = re.compile(r"youtube(?:-nocookie)?\.com/embed/(?P<video_id>[A-Za-z0-9_-]{11})", re.IGNORECASE)
CARD_RE = re.compile(r'<div class="trailers".*?</div>\s*(?=<div class="trailers"|\Z)', re.IGNORECASE | re.DOTALL)
ALT_RE = re.compile(r'alt="(?P<alt>[^"]*)"', re.IGNORECASE)
LABEL_RE = re.compile(r'<p class="small[^"]*">(?P<label>[^<]*)</p>', re.IGNORECASE)
TAG_RE = re.compile(r"<[^>]+>")
CHANNEL_ID_RE = re.compile(r"^UC[A-Za-z0-9_-]{22}$")
CHALLENGE_MARKERS = ("just a moment", "cf-browser-verification", "challenge-platform")

# Distributor channels. Reuses the hub's existing hint list and extends it.
DISTRIBUTOR_HINTS = {
    "20th century studios", "a24", "amazon mgm studios", "amazon prime video", "amazon studios",
    "apple tv", "bbc film", "bleecker street", "blumhouse", "cj enm", "columbia pictures",
    "constantin film", "criterion collection", "crunchyroll", "dharma productions",
    "disney", "disney plus", "dreamworks animation", "entertainment film distributors",
    "film4", "focus features", "gkids", "hbo", "hbo max", "hulu", "ifc films", "illumination",
    "jiohotstar", "lionsgate", "lionsgate movies", "lucasfilm", "magnolia pictures", "marvel entertainment",
    "marvel studios", "max", "mgm", "mubi", "neon", "netflix", "netflix india", "nickelodeon movies",
    "paramount pictures", "paramount plus", "pathe", "peacock", "pen movies", "pixar",
    "prime video", "reliance entertainment", "roadside attractions", "saban films",
    "samuel goldwyn films", "searchlight pictures", "shochiku", "showbox", "shudder",
    "signature entertainment", "sony pictures", "sony pictures classics", "sony pictures entertainment",
    "star wars", "starz", "studiocanal", "summit entertainment", "sun pictures", "t series",
    "toei", "toho", "tristar pictures", "united artists", "universal pictures", "vertical",
    "vertical entertainment", "walt disney studios", "warner bros", "warner bros pictures",
    "well go usa entertainment", "wild bunch", "working title", "yash raj films", "yrf",
    "zee studios",
}

# Channels that upload trailers but are NOT the distributor.
AGGREGATOR_HINTS = {
    "movieinsider", "movie insider", "joblo", "joblo movie trailers", "joblo trailers",
    "rotten tomatoes trailers", "rotten tomatoes indie", "rotten tomatoes tv",
    "movieclips", "movieclips trailers", "movieclips indie", "fandango movieclips",
    "kinocheck", "kinocheck international", "filmselect trailer", "filmisnow movie trailers",
    "one media", "trailer city", "screen rant", "collider", "ign", "gamespot", "cinemablend",
    "the take", "new trailer buzz", "movie coverage", "new rockstars", "zero media",
    "moviefone", "fresh movie trailers", "coming soon trailers", "flixzone",
}

_ROMAN = {"ii": "2", "iii": "3", "iv": "4", "v": "5", "vi": "6", "vii": "7", "viii": "8", "ix": "9", "x": "10"}
_NUMBER_WORDS = {"one": "1", "two": "2", "three": "3", "four": "4", "five": "5",
                 "six": "6", "seven": "7", "eight": "8", "nine": "9", "ten": "10"}
_APOSTROPHE_RE = re.compile(r"[’ʼ'`´]")
_DIGIT_RE = re.compile(r"\d+")
_LEADING_ARTICLE_RE = re.compile(r"^(the|a|an)\s+")

# A sequel number is one character but decides which film it is. Plain string
# similarity rates "Camp Rock 3" vs "Camp Rock 4" at ~91, above any sane
# auto-accept threshold, so scores are capped when numeric tokens disagree.
SEQUEL_MISMATCH_CAP = 74
TITLE_MATCH_THRESHOLD = 90
TITLE_CANDIDATE_FLOOR = 72


class MovieInsiderBlocked(RuntimeError):
    """MovieInsider served a Cloudflare interstitial instead of the page."""


class ParserBroken(RuntimeError):
    """A listing page returned zero trailers - the site markup probably changed."""


# ---------------------------------------------------------------------------
# Title handling
# ---------------------------------------------------------------------------
def match_key(title: str) -> str:
    """Comparable form of a title.

    Builds on the hub's shared ``normalize_title`` so behaviour stays consistent
    with the IMDb tools, then adds apostrophe removal, roman/word numeral
    folding and leading-article stripping.

    >>> match_key("Dune: Part Two (2024)")
    'dune 2'
    >>> match_key("The Batman")
    'batman'
    """
    text = _APOSTROPHE_RE.sub("", str(title or ""))
    text = normalize_title(text)
    text = re.sub(r"\b(part|chapter|vol|volume|episode)\b", " ", text)
    tokens = [_ROMAN.get(tok, _NUMBER_WORDS.get(tok, tok)) for tok in text.split()]
    text = " ".join(tokens).strip()
    return _LEADING_ARTICLE_RE.sub("", text).strip()


def title_score(left: str, right: str) -> int:
    """0-100 similarity between two already-normalised title keys."""
    if not left or not right:
        return 0
    if left == right:
        return 100

    left_tokens, right_tokens = left.split(), right.split()
    shared = len(set(left_tokens) & set(right_tokens))
    denominator = max(len(set(left_tokens) | set(right_tokens)), 1)
    score = int((shared / denominator) * 100)

    if left in right or right in left:
        score = max(score, 88)

    if sorted(_DIGIT_RE.findall(left)) != sorted(_DIGIT_RE.findall(right)):
        score = min(score, SEQUEL_MISMATCH_CAP)
    return score


def classify_channel(channel_title: str) -> str:
    """``distributor`` | ``aggregator`` | ``unknown`` from the channel name."""
    norm = normalize_title(channel_title or "")
    if not norm:
        return "unknown"
    norm = re.sub(r"\b(official|channel|movies|uk|us|india|brasil|latinoamerica)\b\s*$", "", norm).strip()

    if norm in AGGREGATOR_HINTS:
        return "aggregator"
    if norm in DISTRIBUTOR_HINTS:
        return "distributor"
    for hint in AGGREGATOR_HINTS:
        if len(hint) > 6 and hint in norm:
            return "aggregator"
    for hint in DISTRIBUTOR_HINTS:
        if len(hint) > 6 and hint in norm:
            return "distributor"
    if re.search(r"\b(pictures|studios|films|film|entertainment|productions|distribution)\b", norm):
        return "distributor"
    return "unknown"


def parse_channel_reference(raw: str) -> tuple[str, str]:
    """Classify whatever is in the sheet: ``(kind, value)``.

    The same channel is reachable as ``/channel/UC…``, ``/@handle``, ``/c/name``
    and ``/user/name``. Comparing URL text is the most common source of false
    mismatches in manual review, so everything is reduced to a channel ID before
    comparison.
    """
    text = str(raw or "").strip()
    if not text:
        return "empty", ""
    if CHANNEL_ID_RE.match(text):
        return "id", text
    if text.startswith("@") and "/" not in text:
        return "handle", text.lstrip("@")

    if "youtube.com" in text or "youtu.be" in text or text.startswith("/"):
        path = re.sub(r"^https?://", "", text).split("?")[0].split("#")[0]
        parts = [p for p in path.split("/") if p and "youtu" not in p.lower()]
        for index, part in enumerate(parts):
            nxt = parts[index + 1] if index + 1 < len(parts) else ""
            if part == "channel" and nxt:
                return ("id" if CHANNEL_ID_RE.match(nxt) else "unknown"), nxt
            if part in {"c", "user"} and nxt:
                return "legacy", nxt
            if part.startswith("@") and len(part) > 1:
                return "handle", part.lstrip("@")
        if parts and parts[0] not in {"watch", "playlist", "results", "shorts", "embed"}:
            return "legacy", parts[0]
        return "unknown", text

    if re.match(r"^[A-Za-z0-9_.\- ]{2,80}$", text):
        return "handle", text.replace(" ", "")
    return "unknown", text


# ---------------------------------------------------------------------------
# Input parsing
# ---------------------------------------------------------------------------
class TitleInput:
    def __init__(self, title: str, channel: str = "", year: str = "", imdb_id: str = "",
                 distributor: str = "") -> None:
        self.title = _clean(title)
        self.channel = _clean(channel)
        self.year = _year_from_value(year)
        self.imdb_id = _clean(imdb_id)
        self.distributor = _clean(distributor)


# ---------------------------------------------------------------------------
# Service
# ---------------------------------------------------------------------------
class MovieTrailerChannelService:
    def __init__(
        self,
        http_client: HttpClient,
        api_key: str = "",
        fallback_service: Any | None = None,
        page_budget: int = DEFAULT_PAGE_BUDGET,
        alpha_page_budget: int = DEFAULT_ALPHA_PAGE_BUDGET,
    ) -> None:
        self.http_client = http_client
        self.api_key = (api_key or "").strip()
        self.fallback_service = fallback_service
        self.page_budget = page_budget
        self.alpha_page_budget = alpha_page_budget
        self.quota_units = 0

    # -- public entry point ------------------------------------------------
    def validate_bulk(
        self,
        bulk_text: str = "",
        file_content: bytes | None = None,
        filename: str = "",
        api_key_override: str = "",
        page_budget: int | None = None,
        use_fallback: bool = True,
        progress: ProgressCallback | None = None,
    ) -> dict[str, Any]:
        api_key = (api_key_override or self.api_key).strip()
        if not api_key:
            raise ValueError(
                "YouTube Data API key is missing. Add YOUTUBE_API_KEY to .env or the Render "
                "environment, or paste a key into the form. Enable YouTube Data API v3 in Google Cloud."
            )

        self.quota_units = 0
        budget = page_budget or self.page_budget

        if progress:
            progress(5, "Reading input titles")
        items = self._parse_inputs(bulk_text, file_content, filename)
        if not items:
            raise ValueError("Add at least one movie title, or upload a CSV/XLSX file with a title column.")

        # -- Tier A ---------------------------------------------------------
        if progress:
            progress(10, f"Scanning the {budget * LISTING_PAGE_SIZE} newest MovieInsider trailers")
        trailers = self._sweep_trailer_listing(budget, progress)
        index: dict[str, list[dict[str, Any]]] = {}
        for record in trailers:
            index.setdefault(match_key(record["title"]), []).append(record)

        resolved: list[dict[str, Any]] = []
        unmatched: list[TitleInput] = []
        total = max(len(items), 1)

        for position, item in enumerate(items, start=1):
            if progress and position % 5 == 0:
                progress(45 + int((position / total) * 15), f"Matching {item.title}")
            hit = self._lookup_in_index(index, item)
            if hit:
                resolved.append({"item": item, "record": hit, "source": "MovieInsider trailer list"})
            else:
                unmatched.append(item)

        # -- Tier B ---------------------------------------------------------
        still_missing: list[TitleInput] = []
        for position, item in enumerate(unmatched, start=1):
            if progress:
                progress(60 + int((position / max(len(unmatched), 1)) * 10),
                         f"Searching MovieInsider A-Z for {item.title}")
            record = self._lookup_via_alpha(item)
            if record:
                resolved.append({"item": item, "record": record, "source": "MovieInsider A-Z browse"})
            else:
                still_missing.append(item)

        # -- Resolve channels in bulk ---------------------------------------
        if progress:
            progress(72, "Resolving uploader channels via YouTube videos.list")
        video_ids = [entry["record"]["youtube_video_id"] for entry in resolved if entry["record"].get("youtube_video_id")]
        videos = self._fetch_videos(video_ids, api_key)
        channels = self._fetch_channels({v["channel_id"] for v in videos.values() if v.get("channel_id")}, api_key)

        # -- Canonicalise the sheet's channel references --------------------
        if progress:
            progress(82, "Canonicalising the channel references from your sheet")
        handle_cache: dict[str, str] = {}
        rows: list[dict[str, Any]] = []
        counts: dict[str, int] = {}

        for entry in resolved:
            row = self._build_row(entry, videos, channels, handle_cache, api_key)
            counts[row["Status"]] = counts.get(row["Status"], 0) + 1
            rows.append(row)

        # -- Tier C ---------------------------------------------------------
        fallback_rows = 0
        if still_missing and use_fallback and self.fallback_service is not None:
            if progress:
                progress(88, f"Falling back to YouTube search for {len(still_missing)} unmatched titles")
            for item in still_missing:
                row = self._fallback_row(item, api_key, handle_cache, channels)
                counts[row["Status"]] = counts.get(row["Status"], 0) + 1
                rows.append(row)
                fallback_rows += 1
        else:
            for item in still_missing:
                row = _blank_row(item, TITLE_NOT_FOUND,
                                 "No MovieInsider trailer found within the scanned window.")
                counts[row["Status"]] = counts.get(row["Status"], 0) + 1
                rows.append(row)

        if progress:
            progress(95, "Preparing exports")

        order = [MISMATCH, NEEDS_REVIEW, NOT_OFFICIAL_UPLOAD, TITLE_NOT_FOUND,
                 NO_CHANNEL_FOUND, NO_CHANNEL_PROVIDED, MATCH]
        rows.sort(key=lambda r: (order.index(r["Status"]) if r["Status"] in order else 99, r["Input Title"]))

        summary = (
            f"Checked {len(items)} titles. "
            + " ".join(f"{status}: {counts[status]}." for status in order if counts.get(status))
            + f" Resolved from MovieInsider trailers: {len(resolved)}."
            + (f" Fell back to YouTube search for {fallback_rows}." if fallback_rows else "")
            + f" YouTube quota used: {self.quota_units} units"
            + (f" (the old search-only path would have cost about {len(items) * 101})." if len(items) > 1 else ".")
            + " Official channel means the distributor's YouTube account."
        )

        return {
            "tracker_type": "movie_trailer_channels",
            "title": "Movie Trailer Channel Validation",
            "created_at": utc_now_iso(),
            "source_url": f"{MOVIEINSIDER_BASE}{TRAILERS_PATH}",
            "summary": summary,
            "sections": [
                {
                    "key": "movie_trailer_channels",
                    "title": "Official Distributor YouTube Channels",
                    "columns": MOVIE_TRAILER_CHANNEL_COLUMNS,
                    "rows": rows,
                    "row_count": len(rows),
                    "supports_google": False,
                }
            ],
        }

    # -- Tier A: trailer listing sweep -------------------------------------
    def _sweep_trailer_listing(self, budget: int, progress: ProgressCallback | None) -> list[dict[str, Any]]:
        records: list[dict[str, Any]] = []
        seen: set[str] = set()

        for page_index in range(max(budget, 1)):
            offset = page_index * LISTING_PAGE_SIZE
            url = f"{MOVIEINSIDER_BASE}{TRAILERS_PATH}" + (f"?page_offset={offset}" if offset else "")
            try:
                html = self._get_html(url)
            except MovieInsiderBlocked:
                raise
            except Exception:
                break

            page_records = parse_listing_page(html)
            if not page_records:
                if page_index == 0:
                    raise ParserBroken(
                        "No trailers could be parsed from the first MovieInsider listing page. "
                        "The site markup has probably changed."
                    )
                break

            for record in page_records:
                if record["youtube_video_id"] not in seen:
                    seen.add(record["youtube_video_id"])
                    records.append(record)

            if progress and page_index % 5 == 0:
                progress(10 + int((page_index / max(budget, 1)) * 32),
                         f"Scanned {page_index + 1} trailer pages, {len(records)} trailers")

        return records

    def _lookup_in_index(self, index: dict[str, list[dict[str, Any]]], item: TitleInput) -> dict[str, Any] | None:
        key = match_key(item.title)
        if not key:
            return None
        if key in index:
            return _best_trailer(index[key])

        best, best_score = None, 0
        for candidate_key, candidates in index.items():
            score = title_score(key, candidate_key)
            if score > best_score:
                best, best_score = candidates, score
        if best and best_score >= TITLE_MATCH_THRESHOLD:
            record = dict(_best_trailer(best))
            record["match_score"] = best_score
            return record
        if best and best_score >= TITLE_CANDIDATE_FLOOR:
            record = dict(_best_trailer(best))
            record["match_score"] = best_score
            record["low_confidence"] = True
            return record
        return None

    # -- Tier B: A-Z browse ------------------------------------------------
    def _lookup_via_alpha(self, item: TitleInput) -> dict[str, Any] | None:
        key = match_key(item.title)
        if not key:
            return None
        letter = key[0]
        if not letter.isalnum():
            return None
        if letter.isdigit():
            letter = "0"

        for page_index in range(max(self.alpha_page_budget, 1)):
            offset = page_index * ALPHA_PAGE_SIZE
            url = f"{MOVIEINSIDER_BASE}/alpha/{letter}" + (f"?page_offset={offset}" if offset else "")
            try:
                html = self._get_html(url)
            except MovieInsiderBlocked:
                raise
            except Exception:
                return None

            for movie_id, slug, link_title in parse_alpha_page(html):
                if title_score(key, match_key(link_title)) < TITLE_MATCH_THRESHOLD:
                    continue
                record = self._trailer_from_movie_page(movie_id, slug, link_title)
                if record:
                    return record
        return None

    def _trailer_from_movie_page(self, movie_id: str, slug: str, title: str) -> dict[str, Any] | None:
        try:
            html = self._get_html(f"{MOVIEINSIDER_BASE}/m{movie_id}/{slug}")
        except MovieInsiderBlocked:
            raise
        except Exception:
            return None

        thumb = THUMB_RE.search(html)
        embed = EMBED_RE.search(html)
        video_id = (embed.group("video_id") if embed else None) or (thumb.group("video_id") if thumb else None)
        if not video_id:
            return None

        detail = DETAIL_RE.search(html)
        return {
            "title": title,
            "slug": slug,
            "movie_id": movie_id,
            "video_page_id": detail.group("video_page_id") if detail else "",
            "trailer_label": "Trailer",
            "youtube_video_id": video_id,
            "movieinsider_url": f"{MOVIEINSIDER_BASE}/m{movie_id}/{slug}",
        }

    # -- YouTube API -------------------------------------------------------
    def _fetch_videos(self, video_ids: Iterable[str], api_key: str) -> dict[str, dict[str, str]]:
        unique = list(dict.fromkeys(v for v in video_ids if v))
        output: dict[str, dict[str, str]] = {}

        for start in range(0, len(unique), 50):
            batch = unique[start : start + 50]
            data = self._get_json(YOUTUBE_VIDEOS_URL,
                                  {"part": "snippet", "id": ",".join(batch), "maxResults": "50", "key": api_key},
                                  cost=1)
            for entry in data.get("items", []):
                snippet = entry.get("snippet") or {}
                output[entry.get("id", "")] = {
                    "channel_id": _clean(snippet.get("channelId")),
                    "channel_title": _clean(snippet.get("channelTitle")),
                    "video_title": _clean(snippet.get("title")),
                    "published_at": _clean(snippet.get("publishedAt"))[:10],
                }
        return output

    def _fetch_channels(self, channel_ids: Iterable[str], api_key: str) -> dict[str, dict[str, str]]:
        unique = sorted({c for c in channel_ids if c})
        output: dict[str, dict[str, str]] = {}

        for start in range(0, len(unique), 50):
            batch = unique[start : start + 50]
            data = self._get_json(YOUTUBE_CHANNELS_URL,
                                  {"part": "snippet,statistics", "id": ",".join(batch), "key": api_key},
                                  cost=1)
            for entry in data.get("items", []):
                snippet = entry.get("snippet") or {}
                statistics = entry.get("statistics") or {}
                custom = _clean(snippet.get("customUrl"))
                output[_clean(entry.get("id"))] = {
                    "title": _clean(snippet.get("title")),
                    "handle": custom if custom.startswith("@") else (f"@{custom}" if custom else ""),
                    "subscriber_count": _clean(statistics.get("subscriberCount")),
                    "country": _clean(snippet.get("country")),
                }
        return output

    def _resolve_handle(self, handle: str, api_key: str, cache: dict[str, str]) -> str:
        key = handle.lower()
        if key in cache:
            return cache[key]

        resolved = ""
        for params in ({"forHandle": f"@{handle}"}, {"forUsername": handle}):
            try:
                data = self._get_json(YOUTUBE_CHANNELS_URL,
                                      {"part": "snippet", "key": api_key, **params}, cost=1)
            except Exception:
                continue
            items = data.get("items") or []
            if items:
                resolved = _clean(items[0].get("id"))
                break

        cache[key] = resolved
        return resolved

    # -- Row building ------------------------------------------------------
    def _build_row(self, entry: dict[str, Any], videos: dict[str, dict[str, str]],
                   channels: dict[str, dict[str, str]], handle_cache: dict[str, str],
                   api_key: str) -> dict[str, Any]:
        item: TitleInput = entry["item"]
        record = entry["record"]
        video = videos.get(record.get("youtube_video_id", ""), {})
        channel_id = video.get("channel_id", "")
        channel = channels.get(channel_id, {})
        channel_title = channel.get("title") or video.get("channel_title", "")
        classification = classify_channel(channel_title)

        notes: list[str] = []
        if record.get("low_confidence"):
            notes.append(f"fuzzy title match at {record.get('match_score', 0)}%")

        existing_id, kind = "", "empty"
        if item.channel:
            kind, value = parse_channel_reference(item.channel)
            if kind == "id":
                existing_id = value
            elif kind in {"handle", "legacy"}:
                existing_id = self._resolve_handle(value, api_key, handle_cache)
                if not existing_id:
                    notes.append("could not canonicalise the channel reference in your sheet")

        if not channel_id:
            status = NO_CHANNEL_FOUND
            notes.append("the trailer video is unavailable (deleted, private or region-blocked)")
        elif classification == "aggregator":
            status = NOT_OFFICIAL_UPLOAD
            notes.append("the trailer was uploaded by an aggregator, not the distributor")
        elif kind == "empty":
            status = NO_CHANNEL_PROVIDED
        elif not existing_id:
            status = NEEDS_REVIEW
        elif existing_id == channel_id:
            status = MATCH
        else:
            status = MISMATCH

        if status in {MATCH, MISMATCH} and record.get("low_confidence"):
            status = NEEDS_REVIEW

        if classification == "distributor" and status != NOT_OFFICIAL_UPLOAD:
            notes.append("uploader is a recognised distributor channel")
        elif classification == "unknown" and channel_id:
            notes.append("uploader is not on the distributor or aggregator list")

        video_id = record.get("youtube_video_id", "")
        return {
            "Input Title": item.title,
            "Input Release Year": item.year,
            "IMDb ID": item.imdb_id,
            "Matched MovieInsider Title": record.get("title", ""),
            "Status": status,
            "Official Channel (Distributor)": channel_title,
            "Official Channel ID": channel_id,
            "Official Channel Handle": channel.get("handle", ""),
            "Official Channel URL": _channel_url(channel_id, channel.get("handle", "")),
            "Channel Type": classification,
            "Channel Subscribers": channel.get("subscriber_count", ""),
            "Existing Channel (as provided)": item.channel,
            "Existing Channel ID": existing_id,
            "Trailer": record.get("trailer_label", ""),
            "Trailer URL": f"https://www.youtube.com/watch?v={video_id}" if video_id else "",
            "Trailer Published": video.get("published_at", ""),
            "MovieInsider URL": record.get("movieinsider_url", ""),
            "Resolved Via": entry["source"],
            "Lookup Note": "; ".join(notes),
        }

    def _fallback_row(self, item: TitleInput, api_key: str, handle_cache: dict[str, str],
                      channels: dict[str, dict[str, str]]) -> dict[str, Any]:
        """Tier C: reuse the existing search-based verifier for missing titles."""
        try:
            from app.services.youtube_release_verifier import YouTubeVerifyInput

            legacy_input = YouTubeVerifyInput(
                title=item.title, network=item.distributor, year=str(item.year or "")
            )
            verified = self.fallback_service._verify_one(legacy_input, api_key)
            self.quota_units += 101  # search.list (100) + channels.list (1)
        except Exception as exc:
            return _blank_row(item, TITLE_NOT_FOUND,
                              f"No MovieInsider trailer, and the YouTube search fallback failed: {exc}")

        channel_id = _clean(verified.get("Channel ID"))
        channel_title = _clean(verified.get("YouTube Channel"))
        if not channel_id:
            return _blank_row(item, TITLE_NOT_FOUND,
                              "No MovieInsider trailer, and the YouTube search fallback found no official channel.")

        classification = classify_channel(channel_title)
        notes = [f"resolved by YouTube search fallback (confidence {verified.get('Confidence', '')})",
                 "costs about 101 quota units per title, versus 0.02 via MovieInsider"]

        existing_id, kind = "", "empty"
        if item.channel:
            kind, value = parse_channel_reference(item.channel)
            if kind == "id":
                existing_id = value
            elif kind in {"handle", "legacy"}:
                existing_id = self._resolve_handle(value, api_key, handle_cache)

        if classification == "aggregator":
            status = NOT_OFFICIAL_UPLOAD
        elif kind == "empty":
            status = NO_CHANNEL_PROVIDED
        elif not existing_id:
            status = NEEDS_REVIEW
        elif existing_id == channel_id:
            status = MATCH
        else:
            status = MISMATCH

        # The fallback guesses from search results, so never auto-confirm.
        if status == MATCH and _clean(verified.get("Confirmation")) != "Confirmed":
            status = NEEDS_REVIEW
            notes.append("search fallback was not high-confidence")

        return {
            "Input Title": item.title,
            "Input Release Year": item.year,
            "IMDb ID": item.imdb_id,
            "Matched MovieInsider Title": "",
            "Status": status,
            "Official Channel (Distributor)": channel_title,
            "Official Channel ID": channel_id,
            "Official Channel Handle": "",
            "Official Channel URL": _channel_url(channel_id, ""),
            "Channel Type": classification,
            "Channel Subscribers": "",
            "Existing Channel (as provided)": item.channel,
            "Existing Channel ID": existing_id,
            "Trailer": _clean(verified.get("Video Title")),
            "Trailer URL": _clean(verified.get("YouTube URL")),
            "Trailer Published": _clean(verified.get("YouTube Release Date")),
            "MovieInsider URL": "",
            "Resolved Via": "YouTube search fallback",
            "Lookup Note": "; ".join(notes),
        }

    # -- transport ---------------------------------------------------------
    def _get_html(self, url: str) -> str:
        html = self.http_client.get_text(url)
        head = (html or "")[:4000].lower()
        if any(marker in head for marker in CHALLENGE_MARKERS):
            raise MovieInsiderBlocked(
                "MovieInsider returned a Cloudflare verification page instead of content. "
                "Retry in a minute; if it persists the host IP is being challenged."
            )
        return html

    def _get_json(self, url: str, params: dict[str, str], cost: int) -> dict[str, Any]:
        self.quota_units += cost
        response = self.http_client.session.get(
            url, params=params, timeout=self.http_client.timeout_seconds, allow_redirects=True
        )
        try:
            data = response.json()
        except ValueError:
            data = {}
        if getattr(response, "status_code", 200) >= 400:
            error = data.get("error") if isinstance(data, dict) else {}
            message = _clean(error.get("message") if isinstance(error, dict) else "")
            raise ValueError(f"YouTube Data API request failed: {message or response.status_code}")
        response.raise_for_status()
        return data

    # -- input parsing -----------------------------------------------------
    def _parse_inputs(self, bulk_text: str, file_content: bytes | None, filename: str) -> list[TitleInput]:
        rows: list[TitleInput] = []
        if file_content and filename:
            suffix = Path(filename).suffix.lower()
            if suffix == ".csv":
                rows.extend(_inputs_from_csv(file_content))
            elif suffix in {".xlsx", ".xlsm"}:
                rows.extend(_inputs_from_workbook(file_content))
            else:
                raise ValueError("Upload a CSV, XLSX or XLSM file.")
        rows.extend(_inputs_from_text(bulk_text))
        return [row for row in rows if row.title]


# ---------------------------------------------------------------------------
# Parsers (module level so they are unit-testable without HTTP)
# ---------------------------------------------------------------------------
def parse_listing_page(html: str) -> list[dict[str, Any]]:
    """Parse one /movie-trailers page.

    The ``<h3>`` heading is truncated for long titles ("Children of Blood and
    Bo..."), so the image ``alt`` attribute is the authoritative title source.
    Falls back to a whole-page regex sweep if the card markup changes.
    """
    records: list[dict[str, Any]] = []
    seen: set[str] = set()

    for card in CARD_RE.findall(html or ""):
        thumb = THUMB_RE.search(card)
        if not thumb:
            continue
        detail = DETAIL_RE.search(card)
        alt = ALT_RE.search(card)
        label = LABEL_RE.search(card)

        title = _clean(re.sub(r"\s+trailer\s*$", "", unescape(alt.group("alt")) if alt else "", flags=re.IGNORECASE))
        if not title or title.endswith("..."):
            heading = re.search(r'<h3 class="media-heading">(.*?)</h3>', card, re.DOTALL)
            candidate = _clean(TAG_RE.sub(" ", heading.group(1))) if heading else ""
            title = candidate[:-3].strip() if candidate.endswith("...") else (candidate or title)

        video_id = thumb.group("video_id")
        if video_id in seen:
            continue
        seen.add(video_id)

        records.append(
            {
                "title": title,
                "slug": detail.group("slug") if detail else thumb.group("slug"),
                "movie_id": detail.group("movie_id") if detail else "",
                "video_page_id": detail.group("video_page_id") if detail else "",
                "trailer_label": _clean(label.group("label")) if label else "Trailer",
                "youtube_video_id": video_id,
                "movieinsider_url": (
                    f"{MOVIEINSIDER_BASE}/m{detail.group('movie_id')}/{detail.group('slug')}"
                    f"/videos/{detail.group('video_page_id')}" if detail else ""
                ),
            }
        )

    if records:
        return records

    for thumb in THUMB_RE.finditer(html or ""):
        video_id = thumb.group("video_id")
        if video_id in seen:
            continue
        seen.add(video_id)
        slug = thumb.group("slug")
        records.append(
            {
                "title": slug.replace("-", " ").title(),
                "slug": slug,
                "movie_id": "",
                "video_page_id": "",
                "trailer_label": "Trailer",
                "youtube_video_id": video_id,
                "movieinsider_url": "",
            }
        )
    return records


def parse_alpha_page(html: str) -> list[tuple[str, str, str]]:
    """Parse /alpha/<letter> into ``(movie_id, slug, link_text)`` triples."""
    output: list[tuple[str, str, str]] = []
    seen: set[str] = set()

    for match in re.finditer(
        r'<a[^>]+href="[^"]*?/m(?P<movie_id>\d+)/(?P<slug>[a-z0-9\-]+)"[^>]*>(?P<text>.*?)</a>',
        html or "", re.IGNORECASE | re.DOTALL,
    ):
        movie_id, slug = match.group("movie_id"), match.group("slug")
        if movie_id in seen:
            continue
        text = _clean(TAG_RE.sub(" ", unescape(match.group("text"))))
        if not text:
            continue
        seen.add(movie_id)
        output.append((movie_id, slug, text))
    return output


def _best_trailer(candidates: list[dict[str, Any]]) -> dict[str, Any]:
    """Prefer 'Official Trailer' over teasers; newest wins ties."""

    def rank(record: dict[str, Any]) -> tuple[int, int]:
        label = (record.get("trailer_label") or "").lower()
        if "official trailer" in label:
            tier = 0
        elif "final trailer" in label:
            tier = 1
        elif "trailer" in label:
            tier = 2
        elif "teaser" in label:
            tier = 3
        else:
            tier = 4
        try:
            recency = int(record.get("video_page_id") or 0)
        except (TypeError, ValueError):
            recency = 0
        return (tier, -recency)

    return sorted(candidates, key=rank)[0]


def _channel_url(channel_id: str, handle: str) -> str:
    if handle:
        return f"https://www.youtube.com/{handle}"
    return f"https://www.youtube.com/channel/{channel_id}" if channel_id else ""


def _blank_row(item: TitleInput, status: str, note: str) -> dict[str, Any]:
    return {
        "Input Title": item.title,
        "Input Release Year": item.year,
        "IMDb ID": item.imdb_id,
        "Matched MovieInsider Title": "",
        "Status": status,
        "Official Channel (Distributor)": "",
        "Official Channel ID": "",
        "Official Channel Handle": "",
        "Official Channel URL": "",
        "Channel Type": "",
        "Channel Subscribers": "",
        "Existing Channel (as provided)": item.channel,
        "Existing Channel ID": "",
        "Trailer": "",
        "Trailer URL": "",
        "Trailer Published": "",
        "MovieInsider URL": "",
        "Resolved Via": "",
        "Lookup Note": note,
    }


def _inputs_from_text(raw_text: str) -> list[TitleInput]:
    """One title per line. Optional pipe-separated extras:
    ``Title | channel | year | imdb_id | distributor``"""
    rows = []
    for line in (raw_text or "").splitlines():
        cleaned = line.strip()
        if not cleaned:
            continue
        parts = [part.strip() for part in cleaned.split("|")]
        rows.append(
            TitleInput(
                title=parts[0],
                channel=parts[1] if len(parts) > 1 else "",
                year=parts[2] if len(parts) > 2 else "",
                imdb_id=parts[3] if len(parts) > 3 else "",
                distributor=parts[4] if len(parts) > 4 else "",
            )
        )
    return rows


def _inputs_from_csv(content: bytes) -> list[TitleInput]:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("Could not decode the uploaded CSV file.")

    sample = "\n".join(text.splitlines()[:5])
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        delimiter = ","
    reader = csv.DictReader(StringIO(text), delimiter=delimiter)
    return [_input_from_mapping(row) for row in reader]


def _inputs_from_workbook(content: bytes) -> list[TitleInput]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook.worksheets[0]
    rows = list(worksheet.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        return []

    headers = [_normalize_header(value) for value in rows[0]]
    output = []
    for values in rows[1:]:
        if not any(_clean(value) for value in values):
            continue
        row = {headers[index]: (values[index] if index < len(values) else "") for index in range(len(headers))}
        output.append(_input_from_mapping(row))
    return output


def _input_from_mapping(row: dict[Any, Any]) -> TitleInput:
    return TitleInput(
        title=_first_value(row, ["title", "movie", "movie title", "film", "name", "title name", "input title"]),
        channel=_first_value(row, [
            "current channel url", "current_channel_url", "current channel", "channel url",
            "channel_url", "youtube channel url", "youtube channel", "youtube_channel",
            "channel id", "channel_id", "youtube channel id", "channel", "youtube",
        ]),
        year=_first_value(row, ["release year", "release_year", "year", "release date", "release_date", "date"]),
        imdb_id=_first_value(row, ["imdb id", "imdb_id", "imdb", "imdb code", "tconst", "ttcode"]),
        distributor=_first_value(row, ["distributor", "studio", "network", "network distributor", "publisher"]),
    )


def _first_value(row: dict[Any, Any], keys: list[str]) -> str:
    normalized = {_normalize_header(key): value for key, value in row.items()}
    for key in keys:
        value = normalized.get(_normalize_header(key))
        if _clean(value):
            return _clean(value)
    return ""


def _normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("_", " ").strip().lower())


def _year_from_value(value: Any) -> str:
    match = re.search(r"\b(19\d{2}|20\d{2})\b", _clean(value))
    return match.group(1) if match else ""


def _clean(value: Any) -> str:
    return re.sub(r"\s+", " ", unescape(str(value or "").replace("\xa0", " "))).strip()
