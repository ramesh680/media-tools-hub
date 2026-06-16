from __future__ import annotations

from datetime import datetime
from html import unescape
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Callable
import csv
import re

from openpyxl import load_workbook

from app.models import YOUTUBE_RELEASE_COLUMNS, utc_now_iso
from app.services.http_client import HttpClient
from app.services.imdb import normalize_title


ProgressCallback = Callable[[int, str], None]

YOUTUBE_SEARCH_URL = "https://www.googleapis.com/youtube/v3/search"
YOUTUBE_CHANNELS_URL = "https://www.googleapis.com/youtube/v3/channels"

OFFICIAL_CHANNEL_HINTS = {
    "20th century studios",
    "a24",
    "amazon prime video",
    "amazon studios",
    "apple tv",
    "bbc",
    "bleecker street",
    "criterion collection",
    "dharma productions",
    "disney",
    "disney plus",
    "focus features",
    "hbo",
    "hulu",
    "ifc films",
    "jiohotstar",
    "lionsgate movies",
    "magnolia pictures",
    "max",
    "mgm",
    "mubi",
    "neon",
    "netflix",
    "netflix india",
    "paramount pictures",
    "paramount plus",
    "peacock",
    "prime video",
    "searchlight pictures",
    "shudder",
    "sony pictures",
    "t series",
    "universal pictures",
    "vertical",
    "warner bros",
    "warner bros pictures",
    "yrf",
    "zee studios",
}

TRAILER_TERMS = {
    "official trailer",
    "official teaser",
    "teaser trailer",
    "final trailer",
    "trailer",
    "teaser",
}

NOISE_TERMS = {
    "breakdown",
    "concept trailer",
    "ending explained",
    "fan made",
    "fan trailer",
    "first look reaction",
    "full movie explained",
    "interview",
    "reaction",
    "recap",
    "review",
    "trailer reaction",
}


class YouTubeReleaseVerifierService:
    def __init__(self, http_client: HttpClient, api_key: str = "") -> None:
        self.http_client = http_client
        self.api_key = api_key.strip()

    def verify_bulk(
        self,
        bulk_text: str = "",
        file_content: bytes | None = None,
        filename: str = "",
        api_key_override: str = "",
        progress: ProgressCallback | None = None,
    ) -> dict[str, Any]:
        api_key = (api_key_override or self.api_key).strip()
        if not api_key:
            raise ValueError(
                "YouTube Data API key is missing. Add YOUTUBE_API_KEY to .env or .env.example, "
                "or paste the key into the YouTube verifier form. Enable YouTube Data API v3 in Google Cloud."
            )
        if progress:
            progress(8, "Reading YouTube verification input")
        input_rows = self._parse_inputs(bulk_text, file_content, filename)
        if not input_rows:
            raise ValueError("Add at least one movie or TV title, or upload a CSV/XLSX file.")

        rows: list[dict[str, Any]] = []
        confirmed = 0
        review = 0
        not_found = 0
        total = max(len(input_rows), 1)
        for index, item in enumerate(input_rows, start=1):
            if progress:
                progress(12 + int((index / total) * 80), f"Checking YouTube for {item.title}")
            output = self._verify_one(item, api_key)
            if output["Confirmation"] == "Confirmed":
                confirmed += 1
            elif output["Confirmation"] == "Review":
                review += 1
            else:
                not_found += 1
            rows.append(output)

        if progress:
            progress(94, "Preparing YouTube verification exports")
        summary = (
            f"Checked {len(input_rows)} titles with YouTube Data API v3. "
            f"Confirmed official-channel trailer matches: {confirmed}. "
            f"Needs review: {review}. Not found: {not_found}. "
            "The YouTube release date is the matched video's published date."
        )
        return {
            "tracker_type": "youtube_release_verifier",
            "title": "YouTube Official Release Verification",
            "created_at": utc_now_iso(),
            "source_url": "https://developers.google.com/youtube/v3/docs/search/list",
            "summary": summary,
            "sections": [
                {
                    "key": "youtube_release_verifier",
                    "title": "Official YouTube Trailer Verification",
                    "columns": YOUTUBE_RELEASE_COLUMNS,
                    "rows": rows,
                    "row_count": len(rows),
                    "supports_google": False,
                }
            ],
        }

    def _parse_inputs(
        self,
        bulk_text: str,
        file_content: bytes | None,
        filename: str,
    ) -> list["YouTubeVerifyInput"]:
        rows: list[YouTubeVerifyInput] = []
        if file_content and filename:
            suffix = Path(filename).suffix.lower()
            if suffix == ".csv":
                rows.extend(_inputs_from_csv(file_content))
            elif suffix in {".xlsx", ".xlsm"}:
                rows.extend(_inputs_from_workbook(file_content))
            else:
                raise ValueError("Upload a CSV, XLSX, or XLSM file for YouTube verification.")
        rows.extend(_inputs_from_text(bulk_text))
        return [row for row in rows if row.title]

    def _verify_one(self, item: "YouTubeVerifyInput", api_key: str) -> dict[str, Any]:
        candidates = self._search_candidates(item, api_key)
        if not candidates:
            return _output_row(
                item,
                confirmation="Not found",
                confidence="0",
                note="No YouTube search results were returned for this title.",
            )
        channel_map = self._fetch_channels({candidate["channel_id"] for candidate in candidates}, api_key)
        scored = [_score_candidate(item, candidate, channel_map.get(candidate["channel_id"], {})) for candidate in candidates]
        scored.sort(key=lambda candidate: candidate["score"], reverse=True)
        best = scored[0]
        if best["confirmation"] == "Not found":
            return _output_row(
                item,
                confirmation="Not found",
                confidence=str(best["score"]),
                note=best["note"],
            )
        return _output_row(
            item,
            confirmation=best["confirmation"],
            confidence=str(best["score"]),
            official_network=best["official_network"],
            channel_title=best["channel_title"],
            channel_id=best["channel_id"],
            video_title=best["video_title"],
            youtube_release_date=best["published_date"],
            youtube_url=f"https://www.youtube.com/watch?v={best['video_id']}",
            matched_keywords=best["matched_keywords"],
            note=best["note"],
        )

    def _search_candidates(self, item: "YouTubeVerifyInput", api_key: str) -> list[dict[str, str]]:
        query = _query_for_item(item)
        data = self._get_json(
            YOUTUBE_SEARCH_URL,
            {
                "part": "snippet",
                "type": "video",
                "maxResults": "10",
                "order": "relevance",
                "regionCode": "US",
                "relevanceLanguage": "en",
                "safeSearch": "none",
                "q": query,
                "key": api_key,
            },
        )
        output: list[dict[str, str]] = []
        seen = set()
        for item_data in data.get("items", []):
            video_id = ((item_data.get("id") or {}).get("videoId") or "").strip()
            snippet = item_data.get("snippet") or {}
            if not video_id or video_id in seen:
                continue
            seen.add(video_id)
            output.append(
                {
                    "video_id": video_id,
                    "video_title": _clean(snippet.get("title")),
                    "description": _clean(snippet.get("description")),
                    "published_date": _date_from_rfc3339(snippet.get("publishedAt")),
                    "channel_id": _clean(snippet.get("channelId")),
                    "channel_title": _clean(snippet.get("channelTitle")),
                }
            )
        return output

    def _fetch_channels(self, channel_ids: set[str], api_key: str) -> dict[str, dict[str, str]]:
        ids = sorted(channel_id for channel_id in channel_ids if channel_id)
        if not ids:
            return {}
        data = self._get_json(
            YOUTUBE_CHANNELS_URL,
            {
                "part": "snippet,statistics",
                "id": ",".join(ids[:50]),
                "key": api_key,
            },
        )
        output: dict[str, dict[str, str]] = {}
        for item in data.get("items", []):
            channel_id = _clean(item.get("id"))
            snippet = item.get("snippet") or {}
            statistics = item.get("statistics") or {}
            output[channel_id] = {
                "title": _clean(snippet.get("title")),
                "description": _clean(snippet.get("description")),
                "custom_url": _clean(snippet.get("customUrl")),
                "subscriber_count": _clean(statistics.get("subscriberCount")),
            }
        return output

    def _get_json(self, url: str, params: dict[str, str]) -> dict[str, Any]:
        response = self.http_client.session.get(
            url,
            params=params,
            timeout=self.http_client.timeout_seconds,
            allow_redirects=True,
        )
        try:
            data = response.json()
        except ValueError:
            data = {}
        if getattr(response, "status_code", 200) >= 400:
            error = data.get("error") if isinstance(data, dict) else {}
            message = _clean(error.get("message") if isinstance(error, dict) else "") or _clean(getattr(response, "text", ""))
            raise ValueError(f"YouTube Data API request failed: {message or response.status_code}")
        response.raise_for_status()
        return data


class YouTubeVerifyInput:
    def __init__(self, title: str, input_type: str = "", network: str = "", year: str = "") -> None:
        self.title = _clean(title)
        self.input_type = _clean(input_type)
        self.network = _clean(network)
        self.year = _year_from_value(year)


def _score_candidate(
    item: YouTubeVerifyInput,
    candidate: dict[str, str],
    channel: dict[str, str],
) -> dict[str, Any]:
    video_title = candidate["video_title"]
    video_text = f"{video_title} {candidate['description']}"
    channel_title = channel.get("title") or candidate["channel_title"]
    channel_text = f"{channel_title} {channel.get('custom_url', '')} {channel.get('description', '')}"

    title_score, title_terms = _title_match_score(item.title, video_title)
    trailer_score, trailer_terms = _trailer_score(video_text)
    official_score, official_terms, official_network = _official_channel_score(item.network, channel_text, channel_title)
    year_score = 5 if item.year and item.year in video_text else 0
    penalty = _noise_penalty(video_text)
    score = max(0, min(100, title_score + trailer_score + official_score + year_score - penalty))

    matched_terms = title_terms + trailer_terms + official_terms
    if item.year and year_score:
        matched_terms.append(f"year:{item.year}")
    if penalty:
        matched_terms.append("noise penalty")

    if score >= 75 and title_score >= 18 and trailer_score >= 12 and official_score >= 25:
        confirmation = "Confirmed"
        note = "High-confidence match to an official or distributor/network YouTube channel."
    elif score >= 55 and title_score >= 18 and trailer_score >= 10:
        confirmation = "Review"
        note = "Potential match found, but official-channel confidence is not high enough for automatic confirmation."
    else:
        confirmation = "Not found"
        note = "No high-confidence official-channel trailer match was found in the top YouTube search results."

    return {
        "score": score,
        "confirmation": confirmation,
        "note": note,
        "official_network": official_network or channel_title,
        "channel_title": channel_title,
        "channel_id": candidate["channel_id"],
        "video_id": candidate["video_id"],
        "video_title": video_title,
        "published_date": candidate["published_date"],
        "matched_keywords": "; ".join(dict.fromkeys(matched_terms)),
    }


def _inputs_from_text(raw_text: str) -> list[YouTubeVerifyInput]:
    rows = []
    for line in raw_text.splitlines():
        cleaned = line.strip()
        if not cleaned:
            continue
        parts = [part.strip() for part in cleaned.split("|")]
        rows.append(
            YouTubeVerifyInput(
                title=parts[0] if len(parts) >= 1 else "",
                input_type=parts[1] if len(parts) >= 2 else "",
                network=parts[2] if len(parts) >= 3 else "",
                year=parts[3] if len(parts) >= 4 else "",
            )
        )
    return rows


def _inputs_from_csv(content: bytes) -> list[YouTubeVerifyInput]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(StringIO(text))
    return [_input_from_mapping(row) for row in reader]


def _inputs_from_workbook(content: bytes) -> list[YouTubeVerifyInput]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook.worksheets[0]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_normalize_header(value) for value in rows[0]]
    output = []
    for values in rows[1:]:
        if not any(_clean(value) for value in values):
            continue
        row = {headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))}
        output.append(_input_from_mapping(row))
    return output


def _input_from_mapping(row: dict[Any, Any]) -> YouTubeVerifyInput:
    return YouTubeVerifyInput(
        title=_first_value(row, ["title", "title name", "name", "movie", "show", "input title"]),
        input_type=_first_value(row, ["type", "title_category", "title category", "category", "input type"]),
        network=_first_value(
            row,
            [
                "network",
                "network distributor",
                "network_distributor",
                "availability network",
                "availability / network",
                "distributor",
                "studio",
                "publisher",
                "company",
                "companies",
            ],
        ),
        year=_first_value(row, ["year", "release year", "release_year", "release date", "release_date", "date"]),
    )


def _first_value(row: dict[Any, Any], keys: list[str]) -> str:
    normalized = {_normalize_header(key): value for key, value in row.items()}
    for key in keys:
        value = normalized.get(_normalize_header(key))
        if _clean(value):
            return _clean(value)
    return ""


def _query_for_item(item: YouTubeVerifyInput) -> str:
    parts = [item.title, "official trailer"]
    if item.network:
        parts.append(item.network)
    if item.year:
        parts.append(item.year)
    return " ".join(parts)


def _title_match_score(input_title: str, video_title: str) -> tuple[int, list[str]]:
    input_norm = normalize_title(input_title)
    video_norm = normalize_title(video_title)
    if not input_norm or not video_norm:
        return 0, []
    if input_norm in video_norm:
        return 35, ["title exact"]
    input_tokens = _important_tokens(input_norm)
    video_tokens = set(_important_tokens(video_norm))
    if not input_tokens:
        return 0, []
    overlap = len([token for token in input_tokens if token in video_tokens])
    ratio = overlap / len(input_tokens)
    return int(ratio * 32), [f"title tokens:{overlap}/{len(input_tokens)}"] if overlap else []


def _trailer_score(text: str) -> tuple[int, list[str]]:
    normalized = normalize_title(text)
    if "official trailer" in normalized:
        return 25, ["official trailer"]
    for term in TRAILER_TERMS:
        if normalize_title(term) in normalized:
            return 18, [term]
    return 0, []


def _official_channel_score(network: str, channel_text: str, channel_title: str) -> tuple[int, list[str], str]:
    normalized_channel = normalize_title(channel_text)
    normalized_title = normalize_title(channel_title)
    terms: list[str] = []
    official_network = ""
    score = 0
    if network:
        network_score = _network_score(network, channel_text)
        if network_score:
            score += network_score
            terms.append("network/distributor channel match")
            official_network = channel_title or network
    for hint in OFFICIAL_CHANNEL_HINTS:
        hint_norm = normalize_title(hint)
        if hint_norm and (hint_norm in normalized_channel or hint_norm in normalized_title):
            score += 30
            terms.append(f"official channel hint:{hint}")
            official_network = channel_title or hint
            break
    if "official" in normalized_channel:
        score += 12
        terms.append("official keyword")
    return min(score, 45), terms, official_network


def _network_score(network: str, channel_text: str) -> int:
    network_tokens = set(_important_tokens(normalize_title(network)))
    channel_tokens = set(_important_tokens(normalize_title(channel_text)))
    if not network_tokens or not channel_tokens:
        return 0
    overlap = len(network_tokens & channel_tokens)
    if overlap == len(network_tokens):
        return 35
    if overlap:
        return 18
    return 0


def _noise_penalty(text: str) -> int:
    normalized = normalize_title(text)
    return 25 if any(normalize_title(term) in normalized for term in NOISE_TERMS) else 0


def _important_tokens(normalized_text: str) -> list[str]:
    stop_words = {"a", "an", "and", "the", "of", "for", "to", "in", "on", "new", "official", "trailer", "teaser"}
    return [token for token in normalized_text.split() if token not in stop_words and len(token) > 1]


def _output_row(
    item: YouTubeVerifyInput,
    confirmation: str,
    confidence: str,
    official_network: str = "",
    channel_title: str = "",
    channel_id: str = "",
    video_title: str = "",
    youtube_release_date: str = "",
    youtube_url: str = "",
    matched_keywords: str = "",
    note: str = "",
) -> dict[str, Any]:
    return {
        "Input Title": item.title,
        "Input Type": item.input_type,
        "Input Network / Distributor": item.network,
        "Input Release Year": item.year,
        "Confirmation": confirmation,
        "Confidence": confidence,
        "Official Trailer Network": official_network,
        "YouTube Channel": channel_title,
        "Channel ID": channel_id,
        "Video Title": video_title,
        "YouTube Release Date": youtube_release_date,
        "YouTube URL": youtube_url,
        "Matched Keywords": matched_keywords,
        "Lookup Note": note,
    }


def _date_from_rfc3339(value: Any) -> str:
    text = _clean(value)
    if not text:
        return ""
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date().isoformat()
    except ValueError:
        return text[:10]


def _year_from_value(value: Any) -> str:
    match = re.search(r"\b(19\d{2}|20\d{2})\b", _clean(value))
    return match.group(1) if match else ""


def _normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("_", " ").strip().lower())


def _clean(value: Any) -> str:
    return re.sub(r"\s+", " ", unescape(str(value or "").replace("\xa0", " "))).strip()
