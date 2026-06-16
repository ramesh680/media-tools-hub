from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from io import StringIO
from typing import Any, Callable
from urllib.parse import quote, unquote, urlparse
import csv
import json
import re

from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models import utc_now_iso
from app.services.http_client import HttpClient
from app.services.imdb import IMDbEnrichmentService, normalize_title
from app.services.metacritic import MetacriticParser, TV_PREMIERE_URL, metacritic_url_for_row


METACRITIC_TV_ARCHIVE_URL = "https://www.metacritic.com/news/tv-calendar-archive-of-past-dates/"
WIKIDATA_API_URL = "https://www.wikidata.org/w/api.php"
YOUTUBE_CHANNELS_URL = "https://www.googleapis.com/youtube/v3/channels"
YOUTUBE_SEARCH_URL = "https://www.googleapis.com/youtube/v3/search"

CORE_COLUMNS = [
    "title",
    "title_category",
    "title_sub_category",
    "genre",
    "primary_genre",
    "companies",
    "brand_set",
    "facebook_page",
]

REPORT_COLUMNS = [
    "Sheet",
    "Row",
    "Title",
    "Column",
    "Level",
    "Current Value",
    "Message",
    "Suggestion",
    "Rule",
]
VALIDATED_WORKBOOK_SUMMARY_COLUMNS = [
    "Sheet",
    "Row",
    "Column",
    "Cell",
    "Rule",
    "Message",
    "Value",
]

VALIDATION_SUMMARY_SHEET = "Validation Summary"
ERROR_FILL = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")
ERROR_FONT = Font(color="9C0006")
SUGGESTION_FILL = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")
SUGGESTION_FONT = Font(color="7F6000")

ALLOWED_TITLE_CATEGORIES = [
    "Airlines",
    "Automotive",
    "Car Rental",
    "Consumer Electronics",
    "CPG",
    "Health & Beauty",
    "Education",
    "Energy",
    "Venues, Events & Attractions",
    "Fashion",
    "Financial Services",
    "Food Products",
    "Health, Wellness, Fitness",
    "Hospital & Health Care",
    "Hospitality",
    "Insurance",
    "Internet Services",
    "IT, Internet, Computing",
    "Legal",
    "Government Entities",
    "Marketing, Advertising and Research",
    "Materials and Construction",
    "Media",
    "Movies",
    "Music and Entertainment",
    "TV Network",
    "Non-Profit/Charity/Philanthropy",
    "Pets",
    "Pet Foods & Pet Supplies",
    "Pharmaceuticals",
    "Radio",
    "Real Estate",
    "Restaurants",
    "Retail",
    "Beverages",
    "Sports Franchise",
    "Sports Organizations and Bodies",
    "Film Studio",
    "Supermarket, Grocery, Food & Convenience Stores",
    "Talent",
    "Tourism Boards",
    "Travel",
    "TV Shows",
    "Video Game",
    "Video Game Publishers",
    "Wireless and Telecom",
    "Publishers",
    "Podcasts",
    "Other",
    "Manufacturing & Infrastructure",
]

CATEGORY_LOOKUP = {normalize_title(value): value for value in ALLOWED_TITLE_CATEGORIES}
GENRE_REQUIRED_CATEGORIES = {"Media", "Movies", "TV Shows"}
MOVIE_TV_CATEGORIES = {"Movies", "TV Shows"}
DAR_COMPANY_VALUES = {"Pristine Brand", "Pristine Talent", "Pristine Film"}
DAR_BRAND_SET = "Pristine DAR Brands"
NON_DAR_BRAND_SET = "Competitive View"
INVALID_STRINGS = {"", "#na", "#n/a", "n/a", "na", "nan", "none", "null"}
FACEBOOK_BLOCKED_PATHS = ["/p/", "/php/", "/pages/"]
GENDER_TERMS = {"female", "male", "woman", "man", "women", "men", "non-binary", "nonbinary"}
ProgressCallback = Callable[[int, str], None]

COLUMN_ALIASES = {
    "title": ["title", "title name", "name"],
    "title_category": ["title_category", "title category", "category"],
    "title_sub_category": ["title_sub_category", "title sub category", "subcategory", "sub category"],
    "genre": ["genre", "genres"],
    "primary_genre": ["primary_genre", "primary genre"],
    "companies": ["companies", "company"],
    "brand_set": ["brand_set", "brand set"],
    "facebook_page": ["facebook_page", "facebook page", "facebook url", "facebook_page_url"],
    "release_date": ["release_date", "release date", "start date", "air date"],
    "network": [
        "network",
        "network_distributor",
        "network distributor",
        "availability / network",
        "availability network",
        "distributor",
    ],
    "imdb_id": ["imdb_id", "imdb id", "ttcode", "imdb ttcode"],
    "imdb_url": ["imdb_url", "imdb url", "imdb link", "imdb title url", "imdb profile url"],
    "metacritic_url": ["metacritic_url", "metacritic url", "metacritic", "metacritic link"],
    "youtube_url": ["youtube_url", "youtube url", "youtube", "trailer url"],
    "youtube_channel_username": [
        "youtube_channel_username",
        "youtube channel username",
        "youtube username",
        "youtube channel",
        "youtube_channel",
    ],
    "youtube_channel_company": [
        "youtube_channel_company",
        "youtube channel company",
        "youtube company",
        "youtube_channel_company_url",
        "youtube channel company url",
    ],
    "wikipedia_url": [
        "wikipedia_url",
        "wikipedia url",
        "wikipedia",
        "wiki_url",
        "wiki url",
        "wiki",
    ],
    "twitter_search_terms": ["twitter_search_terms", "twitter search terms"],
    "twitter_search_term_keywords": [
        "twitter_search_term_keywords",
        "twitter search term keywords",
        "twitter_search_keywords",
        "twitter search keywords",
    ],
    "url_managers": ["url_managers", "url managers", "url_manager", "url manager"],
    "instagram_account": [
        "instagram_account",
        "instagram account",
        "instagram_page",
        "instagram page",
        "instagram_url",
        "instagram url",
        "instagram_username",
        "instagram username",
    ],
    "twitter_account": [
        "twitter_account",
        "twitter account",
        "twitter_page",
        "twitter page",
        "twitter_url",
        "twitter url",
        "twitter_username",
        "twitter username",
        "x_twitter",
        "x/twitter",
        "x twitter",
    ],
    "tiktok_account": [
        "tiktok_account",
        "tiktok account",
        "tiktok_page",
        "tiktok page",
        "tiktok_url",
        "tiktok url",
        "tiktok_username",
        "tiktok username",
    ],
    "threads_account": [
        "threads_account",
        "threads account",
        "thread_account",
        "thread account",
        "threads_page",
        "threads page",
        "thread_page",
        "thread page",
        "threads_url",
        "threads url",
        "thread_url",
        "thread url",
        "threads_username",
        "threads username",
        "thread_username",
        "thread username",
    ],
}

URL_MANAGER_PLATFORM_COLUMNS = [
    ("Facebook", "facebook_page"),
    ("YouTube", "youtube_channel_company"),
    ("Instagram", "instagram_account"),
    ("Twitter/X", "twitter_account"),
    ("TikTok", "tiktok_account"),
    ("Threads", "threads_account"),
]
URL_MANAGER_COMPANY_EXCLUSIONS = {normalize_title("Unknown"), normalize_title("Pristine Brand")}
WIKIDATA_CATEGORY_TERMS = {
    "Movies": {normalize_title("film"), normalize_title("movie")},
    "TV Shows": {
        normalize_title("television"),
        normalize_title("television series"),
        normalize_title("television show"),
        normalize_title("tv series"),
    },
    "Talent": {
        normalize_title("actor"),
        normalize_title("actress"),
        normalize_title("artist"),
        normalize_title("athlete"),
        normalize_title("musician"),
        normalize_title("person"),
        normalize_title("singer"),
    },
}

KNOWN_HEADER_KEYS = {
    alias_key
    for aliases in COLUMN_ALIASES.values()
    for alias_key in [re.sub(r"\s+", " ", alias.replace("_", " ").strip().lower()) for alias in aliases]
}

GENRE_ALIASES = {
    "sci fi": "science fiction",
    "sci-fi": "science fiction",
    "rom com": "romance",
    "rom-com": "romance",
    "doc": "documentary",
    "docs": "documentary",
}


@dataclass
class ParsedSheet:
    name: str
    header_row: int
    display_headers: list[str]
    normalized_headers: set[str]
    rows: list[tuple[int, dict[str, Any]]]


@dataclass
class LookupCandidate:
    release_date: str = ""
    release_precision: str = ""
    genres: list[str] | None = None
    network: str = ""
    imdb_id: str = ""
    metacritic_url: str = ""
    source: str = ""


class ExcelValidatorService:
    def __init__(
        self,
        http_client: HttpClient,
        metacritic_parser: MetacriticParser,
        imdb_service: IMDbEnrichmentService,
        tmdb_api_key: str = "",
        tmdb_read_access_token: str = "",
        youtube_api_key: str = "",
    ) -> None:
        self.http_client = http_client
        self.metacritic_parser = metacritic_parser
        self.imdb_service = imdb_service
        self.tmdb_api_key = tmdb_api_key
        self.tmdb_read_access_token = tmdb_read_access_token
        self.youtube_api_key = youtube_api_key.strip()
        self._tmdb_cache: dict[tuple[str, str, str, str], LookupCandidate | None] = {}
        self._imdb_web_cache: dict[str, LookupCandidate | None] = {}
        self._youtube_channel_cache: dict[str, dict[str, Any]] = {}
        self._wikipedia_cache: dict[str, bool] = {}
        self._wikidata_wikipedia_cache: dict[tuple[str, str], dict[str, str] | None] = {}
        self._metacritic_rows: list[dict[str, str]] | None = None
        self._metacritic_error = ""

    def validate_workbook(
        self,
        content: bytes,
        filename: str,
        progress: ProgressCallback | None = None,
    ) -> dict[str, Any]:
        if progress:
            progress(6, "Reading workbook source")
        sheets = _parse_csv_file(content, filename) if filename.lower().endswith(".csv") else _parse_excel_file(content)
        issues: list[dict[str, Any]] = []
        total_rows = 0
        sheet_summaries: list[dict[str, Any]] = []
        validation_total = max(sum(len(sheet.rows) for sheet in sheets), 1)
        validated_rows = 0

        for sheet in sheets:
            if progress:
                progress(14, f"Checking required columns in {sheet.name}")
            total_rows += len(sheet.rows)
            missing_columns = _missing_core_columns(sheet.normalized_headers)
            for missing in missing_columns:
                issues.append(
                    _issue(
                        sheet.name,
                        "",
                        "",
                        missing,
                        "error",
                        "",
                        f"Required column `{missing}` is missing from this sheet.",
                        f"Add a `{missing}` column. Sheet names can be anything, but these column names are required.",
                        "Required columns",
                    )
                )

            sheet_issue_start = len(issues)
            for row_number, row in sheet.rows:
                validated_rows += 1
                if progress:
                    percent = 18 + int((validated_rows / validation_total) * 70)
                    progress(min(percent, 88), f"Validating {sheet.name} row {row_number}")
                self._validate_row(issues, sheet.name, row_number, row)
            sheet_summaries.append(
                {
                    "name": sheet.name,
                    "header_row": sheet.header_row,
                    "rows": len(sheet.rows),
                    "columns": len([header for header in sheet.display_headers if header]),
                    "issues": len(issues) - sheet_issue_start + len(missing_columns),
                    "missing_columns": missing_columns,
                }
            )

        error_count = sum(1 for item in issues if item["Level"] == "error")
        suggestion_count = sum(1 for item in issues if item["Level"] == "suggestion")
        if progress:
            progress(92, "Highlighting failed cells in the workbook")
        validated_workbook_bytes = _build_validated_workbook(content, filename, sheets, issues)
        validated_filename = _validated_workbook_filename(filename)
        if progress:
            progress(96, "Preparing validation report and exports")
        return {
            "filename": filename,
            "validated_filename": validated_filename,
            "validated_workbook_bytes": validated_workbook_bytes,
            "created_at": utc_now_iso(),
            "sheet_count": len(sheets),
            "row_count": total_rows,
            "sheet_summaries": sheet_summaries,
            "issues": issues,
            "error_count": error_count,
            "suggestion_count": suggestion_count,
            "tmdb_enabled": bool(self.tmdb_api_key or self.tmdb_read_access_token),
            "youtube_api_enabled": bool(self.youtube_api_key),
            "metacritic_error": self._metacritic_error,
            "core_columns": CORE_COLUMNS,
            "allowed_categories": ALLOWED_TITLE_CATEGORIES,
            "report_columns": REPORT_COLUMNS,
        }

    def _validate_row(
        self,
        issues: list[dict[str, Any]],
        sheet_name: str,
        row_number: int,
        row: dict[str, Any],
    ) -> None:
        title = _value_for(row, "title")
        category_value = _value_for(row, "title_category")
        category = _canonical_category(category_value)
        title_sub_category = _value_for(row, "title_sub_category")
        genre = _value_for(row, "genre")
        primary_genre = _value_for(row, "primary_genre")
        companies = _value_for(row, "companies")
        brand_set = _value_for(row, "brand_set")
        facebook_page = _value_for(row, "facebook_page")
        release_date = _parse_date_value(_value_for(row, "release_date"))
        network = _value_for(row, "network")
        imdb_id = _value_for(row, "imdb_id")
        imdb_url = _value_for(row, "imdb_url")
        metacritic_url = _value_for(row, "metacritic_url")
        youtube_url = _value_for(row, "youtube_url")
        youtube_channel_username = _value_for(row, "youtube_channel_username")
        youtube_channel_company = _value_for(row, "youtube_channel_company")
        wikipedia_url = _value_for(row, "wikipedia_url")
        twitter_terms = _value_for(row, "twitter_search_terms")
        twitter_keywords = _value_for(row, "twitter_search_term_keywords")
        url_managers = _value_for(row, "url_managers")
        is_dar = _is_dar_title(title)

        if _is_blankish(title):
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    "",
                    "title",
                    "error",
                    title,
                    "Title cannot be blank, #NA, or N/A.",
                    "Add the title before validation can apply title-specific rules.",
                    "Title required",
                )
            )
        elif is_dar and _is_blankish(_dar_title_base(title)):
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    title,
                    "title",
                    "error",
                    title,
                    "DAR title cannot be blank, #NA, or N/A before the ` - DAR` suffix.",
                    "Replace the placeholder with the real title name before ` - DAR`.",
                    "DAR title required",
                )
            )

        if _is_blankish(category_value):
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    title,
                    "title_category",
                    "error",
                    category_value,
                    "title_category cannot be blank, #NA, or N/A.",
                    "Choose one of the approved title_category values.",
                    "Approved category",
                )
            )
        elif not category:
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    title,
                    "title_category",
                    "error",
                    category_value,
                    "title_category is not in the approved category list.",
                    f"Use one of: {', '.join(ALLOWED_TITLE_CATEGORIES)}",
                    "Approved category",
                )
            )

        if _is_blankish(title_sub_category):
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    title,
                    "title_sub_category",
                    "error",
                    title_sub_category,
                    "title_sub_category cannot be blank.",
                    "Add the most specific sub-category available.",
                    "Sub-category required",
                )
            )
        elif category == "Talent" and not _talent_subcategory_complete(title_sub_category):
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    title,
                    "title_sub_category",
                    "error",
                    title_sub_category,
                    "Talent title_sub_category must include gender and profession.",
                    "Use a value such as `Female Actor`, `Male Athlete`, or another gender + profession pairing.",
                    "Talent sub-category",
                )
            )

        candidate = None
        if category in MOVIE_TV_CATEGORIES and not _is_blankish(title):
            candidate = self._lookup_external_candidate(title, category, release_date, network)

        self._validate_genres(issues, sheet_name, row_number, title, category, genre, primary_genre, candidate)
        self._validate_companies(issues, sheet_name, row_number, title, is_dar, companies)
        self._validate_brand_set(issues, sheet_name, row_number, title, is_dar, brand_set)
        self._validate_facebook(issues, sheet_name, row_number, title, facebook_page)
        self._validate_imdb(
            issues,
            sheet_name,
            row_number,
            title,
            category,
            title_sub_category,
            imdb_id,
            imdb_url,
            candidate,
            release_date,
            row,
        )
        self._validate_metacritic_url(issues, sheet_name, row_number, title, category, metacritic_url, candidate, row)
        self._validate_twitter(issues, sheet_name, row_number, title, category, twitter_terms, twitter_keywords)
        self._validate_youtube(issues, sheet_name, row_number, title, youtube_url)
        self._validate_youtube_channel_username(issues, sheet_name, row_number, title, youtube_channel_username)
        self._validate_youtube_channel_company(issues, sheet_name, row_number, title, youtube_channel_company)
        self._validate_wikipedia_url(issues, sheet_name, row_number, title, category, wikipedia_url)
        self._validate_url_managers(issues, sheet_name, row_number, title, row, url_managers)
        self._validate_lookup_suggestions(
            issues,
            sheet_name,
            row_number,
            title,
            release_date,
            network,
            candidate,
        )

    def _validate_genres(
        self,
        issues: list[dict[str, Any]],
        sheet_name: str,
        row_number: int,
        title: str,
        category: str,
        genre: str,
        primary_genre: str,
        candidate: LookupCandidate | None,
    ) -> None:
        candidate_genres = candidate.genres if candidate else []
        if category in GENRE_REQUIRED_CATEGORIES and _is_blankish(genre):
            suggestion = "Add genre for this title_category."
            if candidate_genres:
                suggestion = f"Suggested genres from {candidate.source}: {', '.join(candidate_genres)}"
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    title,
                    "genre",
                    "error",
                    genre,
                    "genre cannot be blank for TV Shows, Movies, or Media.",
                    suggestion,
                    "Genre required",
                )
            )
        elif not _is_blankish(genre) and candidate_genres:
            missing = _missing_genres(genre, candidate_genres)
            if missing:
                issues.append(
                    _issue(
                        sheet_name,
                        row_number,
                        title,
                        "genre",
                        "suggestion",
                        genre,
                        "genre has a value, so it is not an error. Some available genres may be missing.",
                        f"Consider adding: {', '.join(missing)}",
                        "Genre completeness",
                    )
                )

        if not _is_blankish(genre) and _is_blankish(primary_genre):
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    title,
                    "primary_genre",
                    "error",
                    primary_genre,
                    "primary_genre cannot be blank when genre is populated.",
                    "Set primary_genre to the main value from genre.",
                    "Primary genre required",
                )
            )
        elif not _is_blankish(genre) and not _is_blankish(primary_genre):
            genre_values = {_normalize_genre(item) for item in _split_multi_value(genre)}
            if _normalize_genre(primary_genre) not in genre_values:
                issues.append(
                    _issue(
                        sheet_name,
                        row_number,
                        title,
                        "primary_genre",
                        "suggestion",
                        primary_genre,
                        "primary_genre is not one of the listed genre values.",
                        "Choose a primary_genre that is also present in genre.",
                        "Primary genre consistency",
                    )
                )

    def _validate_companies(
        self,
        issues: list[dict[str, Any]],
        sheet_name: str,
        row_number: int,
        title: str,
        is_dar: bool,
        companies: str,
    ) -> None:
        if is_dar:
            if _is_blankish(companies) or not _contains_any(companies, DAR_COMPANY_VALUES):
                issues.append(
                    _issue(
                        sheet_name,
                        row_number,
                        title,
                        "companies",
                        "error",
                        companies,
                        "DAR titles must include Pristine Brand, Pristine Talent, or Pristine Film in companies.",
                        "Add one of: Pristine Brand, Pristine Talent, Pristine Film.",
                        "DAR companies",
                    )
                )
        elif _is_blankish(companies):
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    title,
                    "companies",
                    "error",
                    companies,
                    "companies cannot be blank for non-DAR titles.",
                    "Add the relevant company value.",
                    "Companies required",
                )
            )

    def _validate_brand_set(
        self,
        issues: list[dict[str, Any]],
        sheet_name: str,
        row_number: int,
        title: str,
        is_dar: bool,
        brand_set: str,
    ) -> None:
        if _is_blankish(brand_set):
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    title,
                    "brand_set",
                    "error",
                    brand_set,
                    "brand_set cannot be blank.",
                    f"Use `{DAR_BRAND_SET}` for DAR titles or `{NON_DAR_BRAND_SET}` for non-DAR titles.",
                    "Brand set required",
                )
            )
            return
        if is_dar and not _contains_value(brand_set, DAR_BRAND_SET):
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    title,
                    "brand_set",
                    "error",
                    brand_set,
                    f"DAR titles must include `{DAR_BRAND_SET}` in brand_set.",
                    f"Add `{DAR_BRAND_SET}`. Other brand_set values can stay.",
                    "DAR brand set",
                )
            )
        if not is_dar and not _contains_value(brand_set, NON_DAR_BRAND_SET):
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    title,
                    "brand_set",
                    "error",
                    brand_set,
                    f"Non-DAR titles must include `{NON_DAR_BRAND_SET}` in brand_set.",
                    f"Add `{NON_DAR_BRAND_SET}`.",
                    "Competitive brand set",
                )
            )

    def _validate_facebook(
        self,
        issues: list[dict[str, Any]],
        sheet_name: str,
        row_number: int,
        title: str,
        facebook_page: str,
    ) -> None:
        if _is_blankish(facebook_page):
            return
        for url in _split_url_list(facebook_page):
            parsed = urlparse(url if re.match(r"^https?://", url, flags=re.IGNORECASE) else f"https://{url}")
            host = parsed.netloc.lower()
            path = parsed.path.lower()
            if "facebook.com" not in host:
                issues.append(
                    _issue(
                        sheet_name,
                        row_number,
                        title,
                        "facebook_page",
                        "error",
                        facebook_page,
                        "facebook_page must contain a Facebook URL when populated.",
                        "Use the official facebook.com page URL or leave the cell blank.",
                        "Facebook URL",
                    )
                )
                continue
            if any(blocked in path for blocked in FACEBOOK_BLOCKED_PATHS):
                issues.append(
                    _issue(
                        sheet_name,
                        row_number,
                        title,
                        "facebook_page",
                        "error",
                        facebook_page,
                        "facebook_page URL cannot contain /p/, /php/, or /pages/.",
                        "Replace it with the clean official Facebook account/page URL.",
                        "Facebook URL path",
                    )
                )

    def _validate_imdb(
        self,
        issues: list[dict[str, Any]],
        sheet_name: str,
        row_number: int,
        title: str,
        category: str,
        title_sub_category: str,
        imdb_id: str,
        imdb_url: str,
        candidate: LookupCandidate | None,
        release_date: date | None,
        row: dict[str, Any],
    ) -> None:
        suggested_imdb = candidate.imdb_id if candidate else ""
        if not suggested_imdb and category == "TV Shows":
            suggested_imdb = self._lookup_imdb_tv_ttcode(title, release_date)
        if category == "Talent":
            self._validate_talent_imdb(issues, sheet_name, row_number, title, title_sub_category, imdb_id, imdb_url, row)
            return

        if category not in MOVIE_TV_CATEGORIES:
            return

        id_code = _extract_imdb_code(imdb_id, "tt")
        url_code = _extract_imdb_code(imdb_url, "tt")
        provided_code = id_code or url_code
        current_column = "imdb_id" if id_code or (imdb_id and not url_code) else "imdb_url"
        current_value = imdb_id if current_column == "imdb_id" else imdb_url

        if _is_blankish(imdb_id) and _is_blankish(imdb_url):
            column = "imdb_id" if _row_has_column(row, "imdb_id") else "imdb_url"
            suggestion = f"Suggested imdb_url: https://www.imdb.com/title/{suggested_imdb}/" if suggested_imdb else "No IMDb ttcode suggestion was found from the local IMDb title.basics index."
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    title,
                    column,
                    "error",
                    "",
                    "IMDb ttcode or IMDb title URL cannot be blank for Movies and TV Shows.",
                    suggestion,
                    "IMDb required",
                )
            )
            return

        if not _is_blankish(imdb_id) and not id_code:
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    title,
                    "imdb_id",
                    "error",
                    imdb_id,
                    "imdb_id should contain an IMDb ttcode.",
                    f"Suggested IMDb ttcode: {suggested_imdb}" if suggested_imdb else "Use a value like tt1234567 or an IMDb title URL containing tt1234567.",
                    "IMDb format",
                )
            )
            return

        if not _is_blankish(imdb_url) and not url_code:
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    title,
                    "imdb_url",
                    "error",
                    imdb_url,
                    "imdb_url should be an IMDb title URL containing a ttcode.",
                    f"Suggested IMDb URL: https://www.imdb.com/title/{suggested_imdb}/" if suggested_imdb else "Use a value like https://www.imdb.com/title/tt1234567/.",
                    "IMDb URL format",
                )
            )
            return

        if id_code and url_code and id_code != url_code:
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    title,
                    "imdb_url",
                    "error",
                    imdb_url,
                    "imdb_id and imdb_url contain different IMDb ttcodes.",
                    f"Make both fields use the same ttcode: {id_code}.",
                    "IMDb consistency",
                )
            )
            return

        if suggested_imdb and provided_code and provided_code != suggested_imdb:
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    title,
                    current_column,
                    "error",
                    current_value,
                    "IMDb ttcode does not match the local IMDb title.basics lookup for this title.",
                    f"Suggested IMDb ttcode: {suggested_imdb}; suggested URL: https://www.imdb.com/title/{suggested_imdb}/",
                    "IMDb title.basics lookup",
                )
            )

    def _validate_metacritic_url(
        self,
        issues: list[dict[str, Any]],
        sheet_name: str,
        row_number: int,
        title: str,
        category: str,
        metacritic_url: str,
        candidate: LookupCandidate | None,
        row: dict[str, Any],
    ) -> None:
        if category not in MOVIE_TV_CATEGORIES:
            return

        metacritic_title = _metacritic_title_for_validation(title)
        suggested_url = candidate.metacritic_url if candidate else ""
        if not suggested_url:
            suggested_url = metacritic_url_for_row(
                {
                    "Title Name": metacritic_title,
                    "Release Type": "Movie" if category == "Movies" else "TV Series",
                    "Content Format": "Movie" if category == "Movies" else "TV",
                },
                default_media_type="movie" if category == "Movies" else "tv",
            )

        if _is_blankish(metacritic_url):
            column = "metacritic_url" if _row_has_column(row, "metacritic_url") else "metacritic_url"
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    title,
                    column,
                    "suggestion",
                    "",
                    "Metacritic URL is blank for this movie/TV title.",
                    f"Suggested Metacritic URL based on title: {suggested_url}",
                    "Metacritic URL lookup",
                )
            )
            return

        parsed = urlparse(
            metacritic_url.strip()
            if re.match(r"^https?://", metacritic_url.strip(), flags=re.IGNORECASE)
            else f"https://{metacritic_url.strip()}"
        )
        host = parsed.netloc.lower()
        if not host.endswith("metacritic.com"):
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    title,
                    "metacritic_url",
                    "error",
                    metacritic_url,
                    "metacritic_url must point to metacritic.com.",
                    f"Suggested Metacritic URL based on title: {suggested_url}",
                    "Metacritic URL format",
                )
            )
            return

        page_title = _metacritic_title_from_url(metacritic_url)
        if not page_title:
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    title,
                    "metacritic_url",
                    "error",
                    metacritic_url,
                    "metacritic_url must point to a Metacritic movie or TV title page.",
                    f"Suggested Metacritic URL based on title: {suggested_url}",
                    "Metacritic URL format",
                )
            )
            return

        if (
            not _metacritic_urls_match(metacritic_url, suggested_url)
            and _title_score(metacritic_title, page_title) < 60
        ):
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    title,
                    "metacritic_url",
                    "error",
                    metacritic_url,
                    "metacritic_url title slug does not match the row title.",
                    f"Suggested Metacritic URL based on title: {suggested_url}",
                    "Metacritic title match",
                )
            )

    def _validate_talent_imdb(
        self,
        issues: list[dict[str, Any]],
        sheet_name: str,
        row_number: int,
        title: str,
        title_sub_category: str,
        imdb_id: str,
        imdb_url: str,
        row: dict[str, Any],
    ) -> None:
        if not _row_has_column(row, "imdb_id") and not _row_has_column(row, "imdb_url"):
            return
        suggested_nmcode = self._lookup_imdb_name_nmcode(title, title_sub_category)
        id_code = _extract_imdb_code(imdb_id, "nm")
        url_code = _extract_imdb_code(imdb_url, "nm")
        provided_code = id_code or url_code
        current_column = "imdb_id" if id_code or (imdb_id and not url_code) else "imdb_url"
        current_value = imdb_id if current_column == "imdb_id" else imdb_url

        if _is_blankish(imdb_id) and _is_blankish(imdb_url):
            if suggested_nmcode:
                issues.append(
                    _issue(
                        sheet_name,
                        row_number,
                        title,
                        "imdb_id" if _row_has_column(row, "imdb_id") else "imdb_url",
                        "suggestion",
                        "",
                        "Talent row has no IMDb nmcode.",
                        f"Suggested IMDb nmcode from name.basics: {suggested_nmcode}; suggested URL: https://www.imdb.com/name/{suggested_nmcode}/",
                        "IMDb name.basics lookup",
                    )
                )
            return

        if not _is_blankish(imdb_id) and not id_code:
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    title,
                    "imdb_id",
                    "error",
                    imdb_id,
                    "Talent imdb_id should contain an IMDb nmcode.",
                    f"Suggested IMDb nmcode: {suggested_nmcode}" if suggested_nmcode else "Use a value like nm1234567 or an IMDb name URL containing nm1234567.",
                    "IMDb name format",
                )
            )
            return

        if not _is_blankish(imdb_url) and not url_code:
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    title,
                    "imdb_url",
                    "error",
                    imdb_url,
                    "Talent imdb_url should be an IMDb name URL containing an nmcode.",
                    f"Suggested IMDb URL: https://www.imdb.com/name/{suggested_nmcode}/" if suggested_nmcode else "Use a value like https://www.imdb.com/name/nm1234567/.",
                    "IMDb name URL format",
                )
            )
            return

        if id_code and url_code and id_code != url_code:
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    title,
                    "imdb_url",
                    "error",
                    imdb_url,
                    "Talent imdb_id and imdb_url contain different IMDb nmcodes.",
                    f"Make both fields use the same nmcode: {id_code}.",
                    "IMDb name consistency",
                )
            )
            return

        if suggested_nmcode and provided_code and provided_code != suggested_nmcode:
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    title,
                    current_column,
                    "suggestion",
                    current_value,
                    "Review this Talent nmcode against IMDb name.basics.",
                    f"Suggested IMDb nmcode: {suggested_nmcode}; suggested URL: https://www.imdb.com/name/{suggested_nmcode}/",
                    "IMDb name.basics lookup",
                )
            )

    def _validate_twitter(
        self,
        issues: list[dict[str, Any]],
        sheet_name: str,
        row_number: int,
        title: str,
        category: str,
        twitter_terms: str,
        twitter_keywords: str,
    ) -> None:
        if category not in MOVIE_TV_CATEGORIES:
            return
        if _is_blankish(twitter_terms):
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    title,
                    "twitter_search_terms",
                    "error",
                    twitter_terms,
                    "twitter_search_terms cannot be blank for Movies and TV Shows.",
                    f"Add search terms such as `{title}` and official network/distributor terms.",
                    "Twitter terms required",
                )
            )
        if _is_blankish(twitter_keywords):
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    title,
                    "twitter_search_term_keywords",
                    "error",
                    twitter_keywords,
                    "twitter_search_term_keywords cannot be blank for Movies and TV Shows.",
                    "Add title keywords, network/distributor names, hashtags, cast, or other useful search keywords.",
                    "Twitter keywords required",
                )
            )

    def _validate_youtube(
        self,
        issues: list[dict[str, Any]],
        sheet_name: str,
        row_number: int,
        title: str,
        youtube_url: str,
    ) -> None:
        if _is_blankish(youtube_url) or _valid_youtube_value(youtube_url, title):
            return
        issues.append(
            _issue(
                sheet_name,
                row_number,
                title,
                "youtube_url",
                "error",
                youtube_url,
                "YouTube must be a YouTube URL, or a channel URL followed by a pipe and title name.",
                f"Use a direct YouTube URL, or use `https://www.youtube.com/@OfficialChannel|{title}`.",
                "YouTube URL",
            )
        )

    def _validate_youtube_channel_username(
        self,
        issues: list[dict[str, Any]],
        sheet_name: str,
        row_number: int,
        title: str,
        youtube_channel_username: str,
    ) -> None:
        if _is_blankish(youtube_channel_username) or "%" not in youtube_channel_username:
            return
        issues.append(
            _issue(
                sheet_name,
                row_number,
                title,
                "youtube_channel_username",
                "error",
                youtube_channel_username,
                "youtube_channel_username contains `%` in the URL.",
                "Replace encoded or copied URL text with the clean official YouTube channel URL or username.",
                "YouTube channel URL",
            )
        )

    def _validate_youtube_channel_company(
        self,
        issues: list[dict[str, Any]],
        sheet_name: str,
        row_number: int,
        title: str,
        youtube_channel_company: str,
    ) -> None:
        if _is_blankish(youtube_channel_company):
            return
        if not self.youtube_api_key:
            return

        for raw_value in _split_manager_values(youtube_channel_company):
            reference = _youtube_channel_reference(raw_value)
            if _is_blankish(reference):
                continue
            try:
                result = self._verify_youtube_channel_reference(reference)
            except Exception as exc:
                issues.append(
                    _issue(
                        sheet_name,
                        row_number,
                        title,
                        "youtube_channel_company",
                        "suggestion",
                        raw_value,
                        "YouTube Data API could not validate youtube_channel_company for this row.",
                        f"Check YOUTUBE_API_KEY, API quota, and the channel value. API response: {exc}",
                        "YouTube channel API verification",
                    )
                )
                continue
            if result.get("valid"):
                continue
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    title,
                    "youtube_channel_company",
                    "error",
                    raw_value,
                    "youtube_channel_company could not be verified as an existing YouTube channel.",
                    result.get("suggestion")
                    or "Use a valid YouTube channel URL, @handle, /channel/UC... URL, legacy /user/ URL, or channel id.",
                    "YouTube channel API verification",
                )
            )

    def _verify_youtube_channel_reference(self, reference: str) -> dict[str, Any]:
        parsed = _parse_youtube_channel_reference(reference)
        cache_key = f"{parsed['kind']}:{normalize_title(parsed['value'])}"
        if cache_key in self._youtube_channel_cache:
            return self._youtube_channel_cache[cache_key]

        result: dict[str, Any]
        if parsed["kind"] == "invalid_url":
            result = {
                "valid": False,
                "suggestion": "Use a YouTube channel URL, @handle, /channel/UC... URL, legacy /user/ URL, or channel id.",
            }
        elif parsed["kind"] == "channel_id":
            channel = self._youtube_channel_by_params({"id": parsed["value"]})
            result = _youtube_channel_result(channel, "No YouTube channel was found for this channel id.")
        elif parsed["kind"] == "username":
            channel = self._youtube_channel_by_params({"forUsername": parsed["value"]})
            result = _youtube_channel_result(channel, "No YouTube channel was found for this legacy username.")
        elif parsed["kind"] == "handle":
            channel = self._youtube_channel_by_params({"forHandle": parsed["value"]})
            if not channel:
                channel = self._youtube_channel_from_search(parsed["value"])
            result = _youtube_channel_result(channel, "No YouTube channel was found for this handle.")
        else:
            channel = self._youtube_channel_from_search(parsed["value"])
            result = _youtube_channel_result(channel, "No YouTube channel was found for this channel text.")

        self._youtube_channel_cache[cache_key] = result
        return result

    def _youtube_channel_by_params(self, params: dict[str, str]) -> dict[str, str] | None:
        data = self._youtube_api_get_json(
            YOUTUBE_CHANNELS_URL,
            {
                "part": "snippet",
                "key": self.youtube_api_key,
                **params,
            },
        )
        return _youtube_channel_from_api_items(data.get("items", []))

    def _youtube_channel_from_search(self, query: str) -> dict[str, str] | None:
        data = self._youtube_api_get_json(
            YOUTUBE_SEARCH_URL,
            {
                "part": "snippet",
                "type": "channel",
                "maxResults": "5",
                "q": query,
                "key": self.youtube_api_key,
            },
        )
        channel_ids = [
            _display_value(((item.get("id") or {}).get("channelId") or ""))
            for item in data.get("items", [])
            if _display_value(((item.get("id") or {}).get("channelId") or ""))
        ]
        if not channel_ids:
            return None
        data = self._youtube_api_get_json(
            YOUTUBE_CHANNELS_URL,
            {
                "part": "snippet",
                "id": ",".join(channel_ids[:5]),
                "key": self.youtube_api_key,
            },
        )
        return _best_youtube_channel_match(query, data.get("items", []))

    def _youtube_api_get_json(self, url: str, params: dict[str, str]) -> dict[str, Any]:
        response = self.http_client.session.get(
            url,
            params=params,
            timeout=self.http_client.timeout_seconds,
            allow_redirects=True,
        )
        try:
            data = response.json()
        except ValueError:
            data = {}
        if getattr(response, "status_code", 200) >= 400:
            error = data.get("error") if isinstance(data, dict) else {}
            message = _display_value(error.get("message") if isinstance(error, dict) else "") or _display_value(
                getattr(response, "text", "")
            )
            raise ValueError(message or f"HTTP {response.status_code}")
        response.raise_for_status()
        return data

    def _validate_wikipedia_url(
        self,
        issues: list[dict[str, Any]],
        sheet_name: str,
        row_number: int,
        title: str,
        category: str,
        wikipedia_url: str,
    ) -> None:
        if _is_blankish(wikipedia_url):
            return

        parsed = _parse_wikipedia_url(wikipedia_url)
        if not parsed["is_url"]:
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    title,
                    "wikipedia_url",
                    "error",
                    wikipedia_url,
                    "wikipedia_url must be a full English Wikipedia article URL.",
                    f"Use an English Wikipedia URL such as https://en.wikipedia.org/wiki/{_wikipedia_slug_for_title(title)}.",
                    "Wikipedia URL",
                )
            )
            return
        if parsed["host"] != "en.wikipedia.org":
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    title,
                    "wikipedia_url",
                    "error",
                    wikipedia_url,
                    "Only English Wikipedia URLs are accepted.",
                    f"Replace this with the en.wikipedia.org page for the title, such as https://en.wikipedia.org/wiki/{_wikipedia_slug_for_title(title)}.",
                    "Wikipedia English URL",
                )
            )
            return
        if not parsed["article_title"]:
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    title,
                    "wikipedia_url",
                    "error",
                    wikipedia_url,
                    "wikipedia_url must point to an English Wikipedia article path under /wiki/.",
                    f"Use https://en.wikipedia.org/wiki/{_wikipedia_slug_for_title(title)}.",
                    "Wikipedia article URL",
                )
            )
            return

        expected_title = _dar_title_base(title) if _is_dar_title(title) else _display_value(title)
        wikidata_candidate = self._wikidata_wikipedia_candidate(expected_title, category)
        if wikidata_candidate and wikidata_candidate.get("url"):
            if _wikipedia_urls_match(wikipedia_url, wikidata_candidate["url"]):
                return
            suggestion = f"Use the English Wikipedia URL from Wikidata for `{expected_title}`: {wikidata_candidate['url']}."
            if wikidata_candidate.get("qid"):
                suggestion += f" Wikidata item: https://www.wikidata.org/wiki/{wikidata_candidate['qid']}."
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    title,
                    "wikipedia_url",
                    "error",
                    wikipedia_url,
                    "wikipedia_url does not match the English Wikipedia URL returned by Wikidata for the row title.",
                    suggestion,
                    "Wikidata Wikipedia URL match",
                )
            )
            return

        if not _wikipedia_article_matches_title(parsed["article_title"], expected_title):
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    title,
                    "wikipedia_url",
                    "error",
                    wikipedia_url,
                    "wikipedia_url article title does not match the row title.",
                    f"Use the English Wikipedia page for `{expected_title}`, such as https://en.wikipedia.org/wiki/{_wikipedia_slug_for_title(expected_title)}.",
                    "Wikipedia title match",
                )
            )
            return

        if not self._wikipedia_page_exists(parsed["canonical_url"]):
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    title,
                    "wikipedia_url",
                    "error",
                    wikipedia_url,
                    "The English Wikipedia page could not be verified from wikipedia.org.",
                    "Check that the page exists and use the live English Wikipedia article URL.",
                    "Wikipedia page verification",
                )
            )

    def _wikipedia_page_exists(self, url: str) -> bool:
        if url in self._wikipedia_cache:
            return self._wikipedia_cache[url]
        try:
            self.http_client.get_text(url)
            exists = True
        except Exception:
            exists = False
        self._wikipedia_cache[url] = exists
        return exists

    def _wikidata_wikipedia_candidate(self, title: str, category: str) -> dict[str, str] | None:
        key = (category, normalize_title(title))
        if key in self._wikidata_wikipedia_cache:
            return self._wikidata_wikipedia_cache[key]
        if _is_blankish(title):
            self._wikidata_wikipedia_cache[key] = None
            return None
        try:
            data = self._wikidata_get(
                {
                    "action": "wbsearchentities",
                    "search": title,
                    "language": "en",
                    "uselang": "en",
                    "format": "json",
                    "limit": "8",
                }
            )
            search_items = data.get("search", [])
            qids = [
                _display_value(item.get("id"))
                for item in search_items
                if re.match(r"^Q\d+$", _display_value(item.get("id")))
            ]
            if not qids:
                self._wikidata_wikipedia_cache[key] = None
                return None

            entity_data = self._wikidata_get(
                {
                    "action": "wbgetentities",
                    "ids": "|".join(qids[:8]),
                    "props": "labels|descriptions|sitelinks",
                    "languages": "en",
                    "format": "json",
                }
            )
        except Exception:
            self._wikidata_wikipedia_cache[key] = None
            return None

        entities = entity_data.get("entities") or {}
        search_meta = {
            _display_value(item.get("id")): {
                "label": _display_value(item.get("label")),
                "description": _display_value(item.get("description")),
            }
            for item in search_items
        }
        best_candidate: dict[str, str] | None = None
        best_score = -1
        for qid in qids:
            entity = entities.get(qid) or {}
            if not entity or "missing" in entity:
                continue
            url = _english_wikipedia_url_from_wikidata(entity)
            if not url:
                continue
            label = _display_value(
                (((entity.get("labels") or {}).get("en") or {}).get("value"))
                or search_meta.get(qid, {}).get("label")
            )
            description = _display_value(
                (((entity.get("descriptions") or {}).get("en") or {}).get("value"))
                or search_meta.get(qid, {}).get("description")
            )
            article_title = _english_wikipedia_title_from_wikidata(entity)
            score = _wikidata_wikipedia_score(title, category, label, description, article_title)
            if score > best_score:
                best_score = score
                best_candidate = {
                    "qid": qid,
                    "url": url,
                    "label": label,
                    "description": description,
                    "article_title": article_title,
                }

        if best_score < 60:
            best_candidate = None
        self._wikidata_wikipedia_cache[key] = best_candidate
        return best_candidate

    def _wikidata_get(self, params: dict[str, str]) -> dict[str, Any]:
        response = self.http_client.session.get(
            WIKIDATA_API_URL,
            params=params,
            timeout=self.http_client.timeout_seconds,
            allow_redirects=True,
        )
        response.raise_for_status()
        return response.json()

    def _validate_url_managers(
        self,
        issues: list[dict[str, Any]],
        sheet_name: str,
        row_number: int,
        title: str,
        row: dict[str, Any],
        url_managers: str,
    ) -> None:
        missing_platforms = []
        company_values = _url_manager_company_values(_value_for(row, "companies"))
        missing_company_values = [
            company
            for company in company_values
            if not _manager_contains(url_managers, company)
        ]
        if missing_company_values:
            missing_platforms.append(f"Companies: {', '.join(missing_company_values)}")

        for platform_name, column_name in URL_MANAGER_PLATFORM_COLUMNS:
            platform_value = _value_for(row, column_name)
            if _is_blankish(platform_value):
                continue
            missing_values = [
                item
                for item in _split_manager_values(platform_value)
                if not _manager_contains(url_managers, item)
            ]
            if missing_values:
                missing_platforms.append(f"{platform_name}: {', '.join(missing_values)}")

        if not missing_platforms:
            return
        issues.append(
            _issue(
                sheet_name,
                row_number,
                title,
                "url_managers",
                "error",
                url_managers,
                "url_managers is missing platform accounts that are present elsewhere in this row.",
                "Add these to url_managers: " + " | ".join(missing_platforms),
                "URL managers completeness",
            )
        )

    def _validate_lookup_suggestions(
        self,
        issues: list[dict[str, Any]],
        sheet_name: str,
        row_number: int,
        title: str,
        release_date: date | None,
        network: str,
        candidate: LookupCandidate | None,
    ) -> None:
        if not candidate:
            return
        candidate_release = _parse_date_value(candidate.release_date)
        if candidate.release_precision == "year" and candidate_release:
            if release_date and release_date.year != candidate_release.year:
                issues.append(
                    _issue(
                        sheet_name,
                        row_number,
                        title,
                        "release_date",
                        "error",
                        release_date.isoformat(),
                        f"release_date year does not match {candidate.source}. IMDb title.basics provides a start year, not a full date.",
                        f"IMDb start year: {candidate_release.year}. Verify the exact United States release date manually if needed.",
                        "IMDb release date",
                    )
                )
            elif not release_date:
                issues.append(
                    _issue(
                        sheet_name,
                        row_number,
                        title,
                        "release_date",
                        "suggestion",
                        "",
                        f"{candidate.source} has a start year for this title, but IMDb title.basics does not provide a full release date.",
                        f"IMDb start year: {candidate_release.year}. Verify the exact United States release date manually if needed.",
                        "IMDb release date",
                    )
                )
        elif candidate_release and release_date and candidate_release != release_date:
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    title,
                    "release_date",
                    "error",
                    release_date.isoformat(),
                    f"release_date does not match the IMDb lookup from {candidate.source}.",
                    f"Use the United States release date from IMDb: {candidate_release.isoformat()}",
                    "IMDb release date",
                )
            )
        elif candidate_release and not release_date:
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    title,
                    "release_date",
                    "error",
                    "",
                    f"{candidate.source} has a release date for this title.",
                    f"Use the United States release date from IMDb: {candidate_release.isoformat()}",
                    "IMDb release date",
                )
            )
        if candidate.network and network and not _network_matches(network, candidate.network):
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    title,
                    "network",
                    "suggestion",
                    network,
                    f"Network/distributor differs from {candidate.source}.",
                    f"Suggested network/distributor: {candidate.network}",
                    "IMDb network/distributor",
                )
            )
        elif candidate.network and not network:
            issues.append(
                _issue(
                    sheet_name,
                    row_number,
                    title,
                    "network",
                    "suggestion",
                    "",
                    f"{candidate.source} has network/distributor-style metadata for this title.",
                    f"Suggested network/distributor: {candidate.network}",
                    "IMDb network/distributor",
                )
            )

    def _lookup_external_candidate(
        self,
        title: str,
        category: str,
        release_date: date | None,
        network: str,
    ) -> LookupCandidate | None:
        imdb_candidate = self._lookup_imdb_title_candidate(title, category, release_date)
        metacritic_candidate = self._lookup_metacritic_candidate(title, category, release_date, network)
        return _merge_candidates(imdb_candidate, metacritic_candidate)

    def _lookup_metacritic_candidate(
        self,
        title: str,
        category: str,
        release_date: date | None,
        network: str,
    ) -> LookupCandidate | None:
        metacritic_title = _metacritic_title_for_validation(title)
        if category == "TV Shows":
            candidate = self._lookup_metacritic_tv(metacritic_title, release_date, network)
            if candidate:
                return candidate
        if category in MOVIE_TV_CATEGORIES:
            media_type = "movie" if category == "Movies" else "tv"
            return LookupCandidate(
                metacritic_url=metacritic_url_for_row(
                    {
                        "Title Name": metacritic_title,
                        "Release Type": "Movie" if category == "Movies" else "TV Series",
                        "Content Format": "Movie" if category == "Movies" else "TV",
                    },
                    default_media_type=media_type,
                ),
                source="Metacritic title URL",
            )
        return None

    def _lookup_tmdb(
        self,
        title: str,
        category: str,
        release_date: date | None,
        network: str,
    ) -> LookupCandidate | None:
        if category not in MOVIE_TV_CATEGORIES or not (self.tmdb_api_key or self.tmdb_read_access_token):
            return None
        key = (category, normalize_title(title), release_date.isoformat() if release_date else "", normalize_title(network))
        if key in self._tmdb_cache:
            return self._tmdb_cache[key]

        media_type = "movie" if category == "Movies" else "tv"
        search_url = f"https://api.themoviedb.org/3/search/{media_type}"
        params: dict[str, Any] = {"query": title, "include_adult": "false", "language": "en-US"}
        if self.tmdb_api_key:
            params["api_key"] = self.tmdb_api_key
        if release_date:
            params["year" if media_type == "movie" else "first_air_date_year"] = release_date.year

        try:
            results = self._tmdb_get_json(search_url, params).get("results", [])[:8]
            candidate = self._best_tmdb_candidate(results, media_type, title, release_date, network)
        except Exception:
            candidate = None
        self._tmdb_cache[key] = candidate
        return candidate

    def _best_tmdb_candidate(
        self,
        results: list[dict[str, Any]],
        media_type: str,
        title: str,
        release_date: date | None,
        network: str,
    ) -> LookupCandidate | None:
        best_score = -1
        best_candidate: LookupCandidate | None = None
        for result in results:
            tmdb_id = result.get("id")
            if not tmdb_id:
                continue
            details_params: dict[str, Any] = {
                "append_to_response": "external_ids,release_dates",
                "language": "en-US",
            }
            if self.tmdb_api_key:
                details_params["api_key"] = self.tmdb_api_key
            details = self._tmdb_get_json(f"https://api.themoviedb.org/3/{media_type}/{tmdb_id}", details_params)
            candidate = _candidate_from_tmdb_details(details, media_type)
            score = _title_score(title, details.get("title") or details.get("name") or result.get("title") or result.get("name") or "")
            if release_date and candidate.release_date:
                candidate_date = _parse_date_value(candidate.release_date)
                if candidate_date == release_date:
                    score += 35
                elif candidate_date and candidate_date.year == release_date.year:
                    score += 15
            if network and candidate.network and _network_matches(network, candidate.network):
                score += 25
            if score > best_score:
                best_score = score
                best_candidate = candidate
        return best_candidate if best_score >= 35 else None

    def _tmdb_get_json(self, url: str, params: dict[str, Any]) -> dict[str, Any]:
        headers = {"Accept": "application/json"}
        if self.tmdb_read_access_token:
            headers["Authorization"] = f"Bearer {self.tmdb_read_access_token}"
        response = self.http_client.session.get(
            url,
            params=params,
            headers=headers,
            timeout=self.http_client.timeout_seconds,
            allow_redirects=True,
        )
        response.raise_for_status()
        return response.json()

    def _lookup_metacritic_tv(
        self,
        title: str,
        release_date: date | None,
        network: str,
    ) -> LookupCandidate | None:
        rows = self._get_metacritic_tv_rows()
        best_score = -1
        best_row: dict[str, str] | None = None
        for row in rows:
            score = _title_score(title, row.get("Title Name", ""))
            if release_date and row.get("Release Date"):
                candidate_date = _parse_date_value(row["Release Date"])
                if candidate_date == release_date:
                    score += 35
                elif candidate_date and abs((candidate_date - release_date).days) <= 14:
                    score += 15
            if network and row.get("Availability / Network") and _network_matches(network, row["Availability / Network"]):
                score += 25
            if score > best_score:
                best_score = score
                best_row = row
        if not best_row or best_score < 45:
            return None
        return LookupCandidate(
            release_date=best_row.get("Release Date", ""),
            genres=_split_multi_value(best_row.get("Genre", "")),
            network=best_row.get("Availability / Network", ""),
            metacritic_url=best_row.get("Metacritic URL")
            or metacritic_url_for_row(best_row, default_media_type="tv"),
            source="Metacritic TV calendar/archive",
        )

    def _get_metacritic_tv_rows(self) -> list[dict[str, str]]:
        if self._metacritic_rows is not None:
            return self._metacritic_rows
        rows: list[dict[str, str]] = []
        errors: list[str] = []
        for url in [TV_PREMIERE_URL, METACRITIC_TV_ARCHIVE_URL]:
            try:
                html = self.http_client.get_text(url)
                rows.extend(self.metacritic_parser.parse_tv_calendar(html, today=date.today()))
            except Exception as exc:
                errors.append(f"{url}: {exc}")
        self._metacritic_error = "; ".join(errors)
        self._metacritic_rows = _dedupe_metacritic(rows)
        return self._metacritic_rows

    def _lookup_imdb_tv_ttcode(self, title: str, release_date: date | None) -> str:
        match = self._lookup_imdb_title_metadata(title, "TV Shows", release_date)
        return match.get("imdb_id", "") if match else ""

    def _lookup_imdb_title_candidate(
        self,
        title: str,
        category: str,
        release_date: date | None,
    ) -> LookupCandidate | None:
        if category not in MOVIE_TV_CATEGORIES:
            return None
        match = self._lookup_imdb_title_metadata(title, category, release_date)
        if not match:
            return None
        imdb_id = match.get("imdb_id", "")
        start_year = _imdb_start_year(match.get("start_year", ""))
        basics_candidate = LookupCandidate(
            release_date=f"{start_year}-01-01" if start_year else "",
            release_precision="year" if start_year else "",
            genres=_split_multi_value(match.get("genres", "")),
            network="",
            imdb_id=imdb_id,
            source="IMDb title.basics",
        )
        web_candidate = self._lookup_imdb_web_title_candidate(imdb_id) if imdb_id else None
        return _merge_candidates(web_candidate, basics_candidate)

    def _lookup_imdb_web_title_candidate(self, imdb_id: str) -> LookupCandidate | None:
        if imdb_id in self._imdb_web_cache:
            return self._imdb_web_cache[imdb_id]
        candidate = None
        try:
            html = self.http_client.get_text(f"https://www.imdb.com/title/{imdb_id}/")
            data = _extract_imdb_title_json_ld(html)
            release_date = _jsonld_release_date(data.get("datePublished")) if data else ""
            genres = _jsonld_genre_values(data.get("genre")) if data else []
            network = _jsonld_company_text(data) if data else ""
            if release_date or genres or network:
                candidate = LookupCandidate(
                    release_date=release_date,
                    release_precision="date" if release_date else "",
                    genres=genres,
                    network=network,
                    imdb_id=imdb_id,
                    source="IMDb title page",
                )
        except Exception:
            candidate = None
        self._imdb_web_cache[imdb_id] = candidate
        return candidate

    def _lookup_imdb_title_metadata(self, title: str, category: str, release_date: date | None) -> dict[str, str]:
        try:
            return self.imdb_service.lookup_title(title, category=category, release_date=release_date) or {}
        except Exception:
            return {}

    def _lookup_imdb_name_nmcode(self, title: str, profession_hint: str = "") -> str:
        try:
            match = self.imdb_service.lookup_name(title, profession_hint=profession_hint)
        except Exception:
            return ""
        return match.get("imdb_id", "") if match else ""


def _parse_excel_file(content: bytes) -> list[ParsedSheet]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    return [_parse_sheet(sheet) for sheet in workbook.worksheets if sheet.max_row]


def _parse_csv_file(content: bytes, filename: str) -> list[ParsedSheet]:
    text = content.decode("utf-8-sig")
    reader = csv.reader(StringIO(text))
    rows = list(reader)
    if not rows:
        return [
            ParsedSheet(
                name=filename.rsplit(".", 1)[0] or "CSV",
                header_row=1,
                display_headers=[],
                normalized_headers=set(),
                rows=[],
            )
        ]
    header_index, header_values = _find_csv_header_row(rows)
    display_headers = [_display_value(value) for value in header_values]
    normalized_headers = [_normalize_header(value) for value in header_values]
    parsed_rows: list[tuple[int, dict[str, Any]]] = []
    for row_number, values in enumerate(rows[header_index + 1 :], start=header_index + 2):
        if not any(not _is_blankish(value) for value in values):
            continue
        row = {}
        for index, header in enumerate(normalized_headers):
            if header and header not in row:
                row[header] = values[index] if index < len(values) else ""
        row["__display_headers__"] = display_headers
        parsed_rows.append((row_number, row))
    return [
        ParsedSheet(
            name=filename.rsplit(".", 1)[0] or "CSV",
            header_row=header_index + 1,
            display_headers=display_headers,
            normalized_headers=set(normalized_headers),
            rows=parsed_rows,
        )
    ]


def _parse_sheet(sheet) -> ParsedSheet:
    header_row, header_values = _find_header_row(sheet)
    display_headers = [_display_value(value) for value in header_values]
    normalized_headers = [_normalize_header(value) for value in header_values]
    rows: list[tuple[int, dict[str, Any]]] = []
    for row_number, values in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if not any(not _is_blankish(value) for value in values):
            continue
        row = {}
        for index, header in enumerate(normalized_headers):
            if header and header not in row:
                row[header] = values[index] if index < len(values) else ""
        row["__display_headers__"] = display_headers
        rows.append((row_number, row))
    return ParsedSheet(
        name=sheet.title,
        header_row=header_row,
        display_headers=display_headers,
        normalized_headers=set(normalized_headers),
        rows=rows,
    )


def _build_validated_workbook(
    content: bytes,
    filename: str,
    sheets: list[ParsedSheet],
    issues: list[dict[str, Any]],
) -> bytes:
    workbook = _load_workbook_for_validation_output(content, filename)
    if VALIDATION_SUMMARY_SHEET in workbook.sheetnames:
        del workbook[VALIDATION_SUMMARY_SHEET]

    _highlight_issues(workbook, sheets, issues)
    _append_validation_summary_sheet(workbook, sheets, issues)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _load_workbook_for_validation_output(content: bytes, filename: str) -> Workbook:
    if filename.lower().endswith(".csv"):
        return _csv_to_workbook(content, filename)
    return load_workbook(BytesIO(content))


def _csv_to_workbook(content: bytes, filename: str) -> Workbook:
    text = content.decode("utf-8-sig")
    rows = list(csv.reader(StringIO(text)))
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = _safe_sheet_title(filename.rsplit(".", 1)[0] or "CSV")
    for row in rows:
        worksheet.append(row)
    return workbook


def _highlight_issues(
    workbook: Workbook,
    sheets: list[ParsedSheet],
    issues: list[dict[str, Any]],
) -> None:
    parsed_by_name = {sheet.name: sheet for sheet in sheets}
    for issue in issues:
        sheet_name = _display_value(issue.get("Sheet"))
        row_number = issue.get("Row")
        column_name = _display_value(issue.get("Column"))
        if not sheet_name or not isinstance(row_number, int) or not column_name:
            continue
        if sheet_name not in workbook.sheetnames:
            continue
        parsed_sheet = parsed_by_name.get(sheet_name)
        if not parsed_sheet:
            continue
        column_index = _column_index_for_issue(parsed_sheet, column_name)
        if column_index is None:
            continue

        worksheet = workbook[sheet_name]
        cell = worksheet.cell(row=row_number, column=column_index)
        is_error = issue.get("Level") == "error"
        cell.fill = ERROR_FILL if is_error else SUGGESTION_FILL
        cell.font = ERROR_FONT if is_error else SUGGESTION_FONT
        _append_cell_comment(cell, _issue_comment(issue))


def _column_index_for_issue(sheet: ParsedSheet, column_name: str) -> int | None:
    aliases = COLUMN_ALIASES.get(column_name, [column_name])
    accepted = {_normalize_header(alias) for alias in aliases}
    accepted.add(_normalize_header(column_name))
    for index, header in enumerate(sheet.display_headers, start=1):
        if _normalize_header(header) in accepted:
            return index
    return None


def _append_cell_comment(cell, comment_body: str) -> None:
    if not comment_body:
        return
    if cell.comment and comment_body in cell.comment.text:
        return
    if cell.comment:
        comment_body = f"{cell.comment.text}\n{comment_body}"
    cell.comment = Comment(comment_body[:30000], "Validator")


def _issue_comment(issue: dict[str, Any]) -> str:
    parts = [
        _display_value(issue.get("Rule")),
        _display_value(issue.get("Message")),
        _display_value(issue.get("Suggestion")),
    ]
    return "\n".join(part for part in parts if part)


def _append_validation_summary_sheet(
    workbook: Workbook,
    sheets: list[ParsedSheet],
    issues: list[dict[str, Any]],
) -> None:
    summary = workbook.create_sheet(VALIDATION_SUMMARY_SHEET)
    summary.append(VALIDATED_WORKBOOK_SUMMARY_COLUMNS)

    if issues:
        parsed_by_name = {sheet.name: sheet for sheet in sheets}
        for issue in issues:
            column_letter, cell_reference = _summary_cell_reference(parsed_by_name, issue)
            summary.append(
                [
                    issue.get("Sheet", ""),
                    issue.get("Row", ""),
                    column_letter,
                    cell_reference,
                    _validated_workbook_rule_name(issue),
                    issue.get("Message", ""),
                    issue.get("Current Value", ""),
                ]
            )
    else:
        summary.append(
            [
                "Workbook",
                "",
                "",
                "",
                "validation_complete",
                "No validation issues found.",
                "",
            ]
        )

    for cell in summary[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=False)
    for row in summary.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    summary.freeze_panes = "A2"
    summary.auto_filter.ref = summary.dimensions
    for column_cells in summary.columns:
        max_length = max(len(_display_value(cell.value)) for cell in column_cells)
        summary.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 48)


def _summary_cell_reference(
    parsed_by_name: dict[str, ParsedSheet],
    issue: dict[str, Any],
) -> tuple[str, str]:
    sheet_name = _display_value(issue.get("Sheet"))
    row_number = issue.get("Row")
    column_name = _display_value(issue.get("Column"))
    parsed_sheet = parsed_by_name.get(sheet_name)
    if not parsed_sheet or not isinstance(row_number, int) or not column_name:
        return column_name, ""
    column_index = _column_index_for_issue(parsed_sheet, column_name)
    if column_index is None:
        return column_name, ""
    column_letter = get_column_letter(column_index)
    return column_letter, f"{column_letter}{row_number}"


def _validated_workbook_rule_name(issue: dict[str, Any]) -> str:
    column_name = _display_value(issue.get("Column"))
    message = _display_value(issue.get("Message")).lower()
    rule = _display_value(issue.get("Rule"))
    if column_name in {"title", "title_category", "title_sub_category"} and "cannot be blank" in message:
        return "not_blank_and_not_in"
    if column_name == "title_sub_category" and "talent" in message and "gender" in message:
        return "talent_subcategory_format"
    normalized = re.sub(r"[^a-z0-9]+", "_", rule.lower()).strip("_")
    return normalized or "validation_issue"


def _safe_sheet_title(value: str) -> str:
    title = re.sub(r"[\[\]:*?/\\]", " ", value).strip() or "Sheet"
    return title[:31]


def _validated_workbook_filename(filename: str) -> str:
    stem = filename.rsplit(".", 1)[0].strip() or "validated_workbook"
    safe_stem = re.sub(r"[^A-Za-z0-9_. -]+", "_", stem).strip(" ._") or "validated_workbook"
    return f"{safe_stem}_validated.xlsx"


def _find_csv_header_row(rows: list[list[str]]) -> tuple[int, list[str]]:
    best_index = 0
    best_values: list[str] = rows[0] if rows else []
    best_score = -1
    for index, values in enumerate(rows[:20]):
        normalized_values = {_normalize_header(value) for value in values if not _is_blankish(value)}
        score = len(normalized_values & KNOWN_HEADER_KEYS)
        if score > best_score:
            best_score = score
            best_index = index
            best_values = values
        if score >= 4:
            break
    return best_index, best_values


def _find_header_row(sheet) -> tuple[int, tuple[Any, ...]]:
    best_row_number = 1
    best_values: tuple[Any, ...] = ()
    best_score = -1
    for row_number, values in enumerate(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 20), values_only=True), start=1):
        normalized_values = {_normalize_header(value) for value in values if not _is_blankish(value)}
        score = len(normalized_values & KNOWN_HEADER_KEYS)
        if score > best_score:
            best_score = score
            best_row_number = row_number
            best_values = values
        if score >= 4:
            break
    return best_row_number, best_values


def _missing_core_columns(normalized_headers: set[str]) -> list[str]:
    missing = []
    for column in CORE_COLUMNS:
        aliases = {_normalize_header(alias) for alias in COLUMN_ALIASES[column]}
        if not normalized_headers & aliases:
            missing.append(column)
    return missing


def _extract_imdb_title_json_ld(html: str) -> dict[str, Any]:
    soup = BeautifulSoup(html, "html.parser")
    for script in soup.find_all("script", attrs={"type": "application/ld+json"}):
        raw = script.string or script.get_text()
        if not raw or not raw.strip():
            continue
        try:
            data = json.loads(raw)
        except json.JSONDecodeError:
            continue
        for item in _jsonld_items(data):
            item_type = item.get("@type")
            if item_type in {"Movie", "TVSeries", "TVEpisode"}:
                return item
        for item in _jsonld_items(data):
            if item.get("datePublished") or item.get("genre"):
                return item
    return {}


def _jsonld_items(data: Any) -> list[dict[str, Any]]:
    if isinstance(data, dict):
        items = [data]
        graph = data.get("@graph")
        if isinstance(graph, list):
            items.extend(item for item in graph if isinstance(item, dict))
        return items
    if isinstance(data, list):
        return [item for item in data if isinstance(item, dict)]
    return []


def _jsonld_release_date(value: Any) -> str:
    parsed = _parse_date_value(value)
    return parsed.isoformat() if parsed else ""


def _jsonld_text_values(value: Any) -> list[str]:
    values: list[str] = []
    if isinstance(value, list):
        for item in value:
            values.extend(_jsonld_text_values(item))
    elif isinstance(value, dict):
        text = _display_value(value.get("name") or value.get("url") or value.get("@id"))
        if text:
            values.append(text)
    else:
        text = _display_value(value)
        if text:
            values.append(text)
    output: list[str] = []
    seen = set()
    for value in values:
        key = normalize_title(value)
        if key and key not in seen:
            seen.add(key)
            output.append(value)
    return output


def _jsonld_genre_values(value: Any) -> list[str]:
    values: list[str] = []
    for item in _jsonld_text_values(value):
        values.extend(_split_multi_value(item) or [item])
    output: list[str] = []
    seen = set()
    for value in values:
        key = _normalize_genre(value)
        if key and key not in seen:
            seen.add(key)
            output.append(value)
    return output


def _jsonld_company_text(data: dict[str, Any]) -> str:
    fields = [
        "productionCompany",
        "publisher",
        "provider",
        "sourceOrganization",
        "copyrightHolder",
    ]
    values: list[str] = []
    for field in fields:
        values.extend(_jsonld_text_values(data.get(field)))
    output: list[str] = []
    seen = set()
    for value in values:
        key = normalize_title(value)
        if key and key not in seen:
            seen.add(key)
            output.append(value)
    return "; ".join(output)


def _imdb_start_year(value: Any) -> str:
    match = re.search(r"\b(18|19|20|21)\d{2}\b", _display_value(value))
    return match.group(0) if match else ""


def _candidate_from_tmdb_details(details: dict[str, Any], media_type: str) -> LookupCandidate:
    genres = [genre["name"] for genre in details.get("genres", []) if genre.get("name")]
    if media_type == "movie":
        release_date = _us_movie_release_date(details) or details.get("release_date", "")
        network = "; ".join(
            company.get("name", "") for company in details.get("production_companies", []) if company.get("name")
        )
    else:
        release_date = details.get("first_air_date", "")
        network = "; ".join(item.get("name", "") for item in details.get("networks", []) if item.get("name"))
    external_ids = details.get("external_ids", {}) or {}
    return LookupCandidate(
        release_date=release_date or "",
        release_precision="date" if release_date else "",
        genres=genres,
        network=network,
        imdb_id=external_ids.get("imdb_id") or "",
        source="TMDB",
    )


def _us_movie_release_date(details: dict[str, Any]) -> str:
    release_dates = details.get("release_dates", {}).get("results", [])
    us_block = next((item for item in release_dates if item.get("iso_3166_1") == "US"), None)
    if not us_block:
        return ""
    dates = []
    for item in us_block.get("release_dates", []):
        raw_date = item.get("release_date", "")
        if item.get("type") in {2, 3, 4, 5, 6} and raw_date:
            parsed = _parse_date_value(raw_date[:10])
            if parsed:
                dates.append(parsed)
    return min(dates).isoformat() if dates else ""


def _merge_candidates(
    primary_candidate: LookupCandidate | None,
    secondary_candidate: LookupCandidate | None,
) -> LookupCandidate | None:
    if primary_candidate and secondary_candidate:
        sources = [primary_candidate.source, secondary_candidate.source]
        release_date = primary_candidate.release_date or secondary_candidate.release_date
        return LookupCandidate(
            release_date=release_date,
            release_precision=primary_candidate.release_precision
            if primary_candidate.release_date
            else secondary_candidate.release_precision,
            genres=primary_candidate.genres or secondary_candidate.genres,
            network=primary_candidate.network or secondary_candidate.network,
            imdb_id=secondary_candidate.imdb_id or primary_candidate.imdb_id,
            metacritic_url=primary_candidate.metacritic_url or secondary_candidate.metacritic_url,
            source=" + ".join(source for source in sources if source),
        )
    return primary_candidate or secondary_candidate


def _issue(
    sheet: str,
    row_number: int | str,
    title: str,
    column: str,
    level: str,
    current_value: str,
    message: str,
    suggestion: str,
    rule: str,
) -> dict[str, Any]:
    return {
        "Sheet": sheet,
        "Row": row_number,
        "Title": title,
        "Column": column,
        "Level": level,
        "Current Value": current_value,
        "Message": message,
        "Suggestion": suggestion,
        "Rule": rule,
    }


def _value_for(row: dict[str, Any], canonical_column: str) -> str:
    for alias in COLUMN_ALIASES.get(canonical_column, [canonical_column]):
        value = row.get(_normalize_header(alias))
        if not _is_blankish(value):
            return _display_value(value)
    return ""


def _row_has_column(row: dict[str, Any], canonical_column: str) -> bool:
    return any(_normalize_header(alias) in row for alias in COLUMN_ALIASES.get(canonical_column, [canonical_column]))


def _extract_imdb_code(value: str, prefix: str = "tt") -> str:
    match = re.search(rf"\b{re.escape(prefix)}\d{{7,12}}\b", _display_value(value), flags=re.IGNORECASE)
    return match.group(0).lower() if match else ""


def _canonical_category(value: str) -> str:
    return CATEGORY_LOOKUP.get(normalize_title(value), "")


def _is_dar_title(title: str) -> bool:
    return _display_value(title).lower().endswith(" - dar")


def _dar_title_base(title: str) -> str:
    return re.sub(r"\s+-\s+dar\s*$", "", _display_value(title), flags=re.IGNORECASE).strip()


def _metacritic_title_for_validation(title: str) -> str:
    return _dar_title_base(title) if _is_dar_title(title) else _display_value(title)


def _talent_subcategory_complete(value: str) -> bool:
    normalized = normalize_title(value)
    gender_terms = {normalize_title(item) for item in GENDER_TERMS}
    has_gender = any(term in normalized for term in gender_terms)
    profession_text = normalized
    for term in sorted(gender_terms, key=len, reverse=True):
        profession_text = re.sub(rf"\b{re.escape(term)}\b", " ", profession_text)
    has_profession = bool(re.sub(r"\s+", " ", profession_text).strip())
    return has_gender and has_profession


def _valid_youtube_value(value: str, title: str) -> bool:
    text = _display_value(value)
    if "|" in text:
        channel_url, title_part = [part.strip() for part in text.split("|", 1)]
        if not _is_youtube_url(channel_url) or _is_blankish(title_part):
            return False
        normalized_title = normalize_title(title)
        normalized_part = normalize_title(title_part)
        return normalized_title == normalized_part or normalized_title in normalized_part or normalized_part in normalized_title
    return all(_is_youtube_url(item) for item in _split_url_list(text))


def _is_youtube_url(value: str) -> bool:
    parsed = urlparse(value.strip() if re.match(r"^https?://", value, flags=re.IGNORECASE) else f"https://{value.strip()}")
    host = parsed.netloc.lower()
    return host.endswith("youtube.com") or host.endswith("youtu.be")


def _split_url_list(value: str) -> list[str]:
    return [part.strip() for part in re.split(r"[\n;,]+", _display_value(value)) if part.strip()]


def _split_manager_values(value: str) -> list[str]:
    return [part.strip() for part in re.split(r"[\n;,]+", _display_value(value)) if part.strip()]


def _youtube_channel_reference(value: str) -> str:
    return _display_value(value).split("|", 1)[0].strip()


def _parse_youtube_channel_reference(value: str) -> dict[str, str]:
    text = _youtube_channel_reference(value)
    if not text:
        return {"kind": "", "value": ""}

    if re.match(r"^UC[A-Za-z0-9_-]{20,}$", text):
        return {"kind": "channel_id", "value": text}
    if text.startswith("@"):
        return {"kind": "handle", "value": text}

    looks_like_url = bool(
        re.match(r"^https?://", text, flags=re.IGNORECASE)
        or re.match(r"^(www\.)?(youtube\.com|youtu\.be)/", text, flags=re.IGNORECASE)
    )
    if looks_like_url:
        parsed = urlparse(text if re.match(r"^https?://", text, flags=re.IGNORECASE) else f"https://{text}")
        host = parsed.netloc.lower()
        if not (host.endswith("youtube.com") or host.endswith("youtu.be")):
            return {"kind": "invalid_url", "value": text}
        parts = [part for part in parsed.path.split("/") if part]
        if len(parts) >= 2 and parts[0].lower() == "channel":
            return {"kind": "channel_id", "value": parts[1]}
        if len(parts) >= 2 and parts[0].lower() == "user":
            return {"kind": "username", "value": parts[1]}
        if parts and parts[0].startswith("@"):
            return {"kind": "handle", "value": parts[0]}
        if len(parts) >= 2 and parts[0].lower() in {"c", "handle"}:
            return {"kind": "search", "value": parts[1]}
        if parts:
            return {"kind": "search", "value": parts[-1]}
        return {"kind": "invalid_url", "value": text}

    return {"kind": "search", "value": text}


def _youtube_channel_from_api_items(items: list[dict[str, Any]]) -> dict[str, str] | None:
    if not items:
        return None
    item = items[0]
    snippet = item.get("snippet") or {}
    return {
        "id": _display_value(item.get("id")),
        "title": _display_value(snippet.get("title")),
        "custom_url": _display_value(snippet.get("customUrl")),
    }


def _best_youtube_channel_match(query: str, items: list[dict[str, Any]]) -> dict[str, str] | None:
    channels = []
    query_token = _normalize_youtube_channel_token(query)
    for item in items:
        snippet = item.get("snippet") or {}
        channel = {
            "id": _display_value(item.get("id")),
            "title": _display_value(snippet.get("title")),
            "custom_url": _display_value(snippet.get("customUrl")),
        }
        channels.append(channel)
        candidates = {
            _normalize_youtube_channel_token(channel["title"]),
            _normalize_youtube_channel_token(channel["custom_url"]),
        }
        if query_token and query_token in candidates:
            return channel
    return channels[0] if channels else None


def _youtube_channel_result(channel: dict[str, str] | None, missing_message: str) -> dict[str, Any]:
    if channel:
        return {
            "valid": True,
            "channel": channel,
            "suggestion": f"Verified YouTube channel: {channel.get('title') or channel.get('id')}.",
        }
    return {
        "valid": False,
        "suggestion": f"{missing_message} Use the official YouTube channel URL, @handle, or /channel/UC... URL.",
    }


def _normalize_youtube_channel_token(value: str) -> str:
    text = _display_value(value).lower().strip()
    if text.startswith("@"):
        text = text[1:]
    parsed = urlparse(text if re.match(r"^https?://", text, flags=re.IGNORECASE) else f"https://{text}")
    if parsed.netloc and parsed.netloc.endswith("youtube.com"):
        parts = [part for part in parsed.path.split("/") if part]
        if parts:
            text = parts[-1]
    return normalize_title(text.lstrip("@"))


def _parse_wikipedia_url(value: str) -> dict[str, str | bool]:
    text = _display_value(value).strip()
    is_url = bool(re.match(r"^https?://", text, flags=re.IGNORECASE))
    if not is_url:
        return {"is_url": False, "host": "", "article_title": "", "canonical_url": ""}
    parsed = urlparse(text)
    host = parsed.netloc.lower()
    if host.startswith("www."):
        host = host[4:]
    path_parts = [part for part in parsed.path.split("/") if part]
    article_title = ""
    if len(path_parts) >= 2 and path_parts[0] == "wiki":
        article_title = unquote(path_parts[1]).replace("_", " ").strip()
    canonical_url = ""
    if host == "en.wikipedia.org" and article_title:
        canonical_url = f"https://en.wikipedia.org/wiki/{path_parts[1]}"
    return {
        "is_url": True,
        "host": host,
        "article_title": article_title,
        "canonical_url": canonical_url,
    }


def _wikipedia_article_matches_title(article_title: str, title: str) -> bool:
    return bool(normalize_title(article_title) and normalize_title(article_title) == normalize_title(title))


def _wikipedia_slug_for_title(title: str) -> str:
    display = _dar_title_base(title) if _is_dar_title(title) else _display_value(title)
    cleaned = re.sub(r"\s+", "_", display.strip())
    return cleaned or "Title"


def _english_wikipedia_title_from_wikidata(entity: dict[str, Any]) -> str:
    return _display_value((((entity.get("sitelinks") or {}).get("enwiki") or {}).get("title")))


def _english_wikipedia_url_from_wikidata(entity: dict[str, Any]) -> str:
    title = _english_wikipedia_title_from_wikidata(entity)
    if not title:
        return ""
    return f"https://en.wikipedia.org/wiki/{quote(title.replace(' ', '_'), safe='()_,-')}"


def _wikipedia_urls_match(left: str, right: str) -> bool:
    return bool(_wikipedia_url_key(left) and _wikipedia_url_key(left) == _wikipedia_url_key(right))


def _wikipedia_url_key(value: str) -> str:
    text = _display_value(value).strip()
    if not text:
        return ""
    parsed = urlparse(text if re.match(r"^https?://", text, flags=re.IGNORECASE) else f"https://{text}")
    host = re.sub(r"^www\.", "", parsed.netloc.lower())
    parts = [part for part in parsed.path.split("/") if part]
    if host != "en.wikipedia.org" or len(parts) < 2 or parts[0].lower() != "wiki":
        return ""
    title = unquote(parts[1]).replace("_", " ")
    return re.sub(r"\s+", " ", title).strip().casefold()


def _wikidata_wikipedia_score(
    title: str,
    category: str,
    label: str,
    description: str,
    article_title: str,
) -> int:
    score = max(_title_score(title, label), _title_score(title, article_title))
    year = _year_from_text(title)
    combined = normalize_title(" ".join([label, description, article_title]))
    if year and year in combined:
        score += 30
    category_terms = WIKIDATA_CATEGORY_TERMS.get(category, set())
    if category_terms and any(term in combined for term in category_terms):
        score += 20
    if article_title:
        score += 5
    return score


def _year_from_text(value: str) -> str:
    match = re.search(r"\b(18|19|20|21)\d{2}\b", _display_value(value))
    return match.group(0) if match else ""


def _metacritic_title_from_url(value: str) -> str:
    text = _display_value(value).strip()
    parsed = urlparse(text if re.match(r"^https?://", text, flags=re.IGNORECASE) else f"https://{text}")
    parts = [part for part in parsed.path.split("/") if part]
    if len(parts) >= 2 and parts[0].lower() in {"movie", "tv"}:
        return unquote(parts[1]).replace("-", " ").strip()
    return ""


def _metacritic_urls_match(left: str, right: str) -> bool:
    return bool(_metacritic_url_key(left) and _metacritic_url_key(left) == _metacritic_url_key(right))


def _metacritic_url_key(value: str) -> str:
    text = _display_value(value).strip()
    if not text:
        return ""
    parsed = urlparse(text if re.match(r"^https?://", text, flags=re.IGNORECASE) else f"https://{text}")
    host = re.sub(r"^www\.", "", parsed.netloc.lower())
    if not host.endswith("metacritic.com"):
        return ""
    parts = [unquote(part).lower() for part in parsed.path.split("/") if part]
    if len(parts) < 2 or parts[0] not in {"movie", "tv"}:
        return ""
    return f"{parts[0]}/{parts[1].strip('/')}"


def _url_manager_company_values(companies: str) -> list[str]:
    values = []
    seen = set()
    for item in _split_multi_value(companies):
        normalized = normalize_title(item)
        if _is_blankish(item) or normalized in URL_MANAGER_COMPANY_EXCLUSIONS:
            continue
        if normalized in seen:
            continue
        values.append(item.strip())
        seen.add(normalized)
    return values


def _manager_contains(url_managers: str, platform_value: str) -> bool:
    manager_text = _display_value(url_managers)
    if _is_blankish(manager_text):
        return False

    manager_normalized = _normalize_manager_text(manager_text)
    value_normalized = _normalize_manager_text(platform_value)
    if value_normalized and value_normalized in manager_normalized:
        return True

    handle = _extract_social_handle(platform_value)
    return bool(handle and handle in manager_normalized)


def _normalize_manager_text(value: str) -> str:
    pieces = [_normalize_manager_piece(part) for part in _split_manager_values(value)]
    if not pieces:
        pieces = [_normalize_manager_piece(value)]
    return " ".join(piece for piece in pieces if piece)


def _normalize_manager_piece(value: str) -> str:
    text = _display_value(value).lower().strip()
    if not text:
        return ""
    parsed = urlparse(text if re.match(r"^https?://", text, flags=re.IGNORECASE) else f"https://{text}")
    if parsed.netloc and "." in parsed.netloc:
        host = re.sub(r"^www\.", "", parsed.netloc.lower())
        path = parsed.path.rstrip("/").lower()
        return f"{host}{path}"
    return text.strip().rstrip("/").lstrip("@")


def _extract_social_handle(value: str) -> str:
    text = _display_value(value).strip()
    parsed = urlparse(text if re.match(r"^https?://", text, flags=re.IGNORECASE) else f"https://{text}")
    path_parts = [part for part in parsed.path.split("/") if part]
    if path_parts:
        return path_parts[-1].lower().lstrip("@")
    if text.startswith("@"):
        return text.lower().lstrip("@")
    return ""


def _missing_genres(existing: str, candidate_genres: list[str]) -> list[str]:
    existing_set = {_normalize_genre(item) for item in _split_multi_value(existing)}
    return [genre for genre in candidate_genres if _normalize_genre(genre) not in existing_set]


def _split_multi_value(value: str) -> list[str]:
    return [part.strip() for part in re.split(r"[,;/|]+", value or "") if part.strip()]


def _normalize_genre(value: str) -> str:
    normalized = normalize_title(value)
    return GENRE_ALIASES.get(normalized, normalized)


def _contains_any(value: str, expected_values: set[str]) -> bool:
    return any(_contains_value(value, expected) for expected in expected_values)


def _contains_value(value: str, expected: str) -> bool:
    normalized_value = normalize_title(value)
    normalized_expected = normalize_title(expected)
    return normalized_expected in normalized_value


def _network_matches(existing: str, candidate: str) -> bool:
    existing_tokens = set(normalize_title(existing).split())
    candidate_tokens = set(normalize_title(candidate).split())
    if not existing_tokens or not candidate_tokens:
        return False
    return bool(existing_tokens & candidate_tokens)


def _title_score(left: str, right: str) -> int:
    left_norm = normalize_title(left)
    right_norm = normalize_title(right)
    if not left_norm or not right_norm:
        return 0
    if left_norm == right_norm:
        return 70
    if left_norm in right_norm or right_norm in left_norm:
        return 45
    left_tokens = set(left_norm.split())
    right_tokens = set(right_norm.split())
    overlap = len(left_tokens & right_tokens)
    return int((overlap / max(len(left_tokens), len(right_tokens))) * 40)


def _parse_date_value(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _display_value(value)
    if _is_blankish(text):
        return None
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        pass
    for fmt in ["%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"]:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("_", " ").strip().lower())


def _display_value(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()


def _is_blankish(value: Any) -> bool:
    normalized = _display_value(value).strip().lower()
    normalized = normalized.replace(" ", "")
    return normalized in INVALID_STRINGS


def _dedupe_metacritic(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    seen = set()
    output = []
    for row in rows:
        key = (normalize_title(row.get("Title Name", "")), row.get("Release Date", ""))
        if key in seen or not key[0]:
            continue
        seen.add(key)
        output.append(row)
    return output
